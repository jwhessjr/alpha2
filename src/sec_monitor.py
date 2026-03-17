"""
sec_monitor.py — Monitor SEC EDGAR for new 10-K and 10-Q filings
for every stock in valuation.db.

Usage:
    python sec_monitor.py               # check and report new filings
    python sec_monitor.py --report      # also write an xlsx summary
    python sec_monitor.py --init        # seed tracker without reporting (first run)

How it works:
  1. Read all (ticker, cik) pairs from valuation.db where cik is populated.
  2. For each CIK, fetch the SEC EDGAR submissions JSON.
  3. Compare the most recent 10-K and 10-Q filing dates to what is stored
     in the filing_tracker table.
  4. Print any new filings to the console; optionally write an xlsx report.
  5. Update filing_tracker so the next run only reports truly new filings.

SEC EDGAR rate limit: 10 requests / second.  We pace at ~5 req/s to be safe.
"""

import sqlite3
import time
import sys
import os
from datetime import date, datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

DB_PATH = "/Volumes/Financial_Data/valuation.db"
OUTPUT_DIR = "/Users/jhess/HessGrp/data"
EDGAR_UA = "AlphaResearch research@example.com"  # SEC requires name + email
FORMS = ("10-K", "10-Q")
SLEEP_BETWEEN_REQUESTS = 0.20  # 5 req/s

EDGAR_SUBMISSIONS = "https://data.sec.gov/submissions/CIK{cik}.json"
EDGAR_FILING_URL = (
    "https://www.sec.gov/cgi-bin/browse-edgar"
    "?action=getcompany&CIK={cik}&type={form}&dateb=&owner=include&count=5"
)

# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

TRACKER_SCHEMA = """
CREATE TABLE IF NOT EXISTS filing_tracker (
    cik          TEXT NOT NULL,
    ticker       TEXT NOT NULL,
    form_type    TEXT NOT NULL,
    last_filing_date   TEXT NOT NULL,
    accession_number   TEXT NOT NULL,
    checked_at   TEXT NOT NULL,
    PRIMARY KEY (cik, form_type)
);
"""


def open_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute(TRACKER_SCHEMA)
    conn.commit()
    return conn


def get_tracked_stocks(conn: sqlite3.Connection) -> list[dict]:
    """Return all (ticker, cik) pairs that have a CIK populated."""
    rows = conn.execute(
        "SELECT ticker, cik, ent_name FROM valuation WHERE cik <> '' ORDER BY ticker"
    ).fetchall()
    return [{"ticker": r[0], "cik": r[1], "name": r[2]} for r in rows]


def get_known_filing(conn: sqlite3.Connection, cik: str, form_type: str) -> dict | None:
    row = conn.execute(
        "SELECT last_filing_date, accession_number FROM filing_tracker "
        "WHERE cik = ? AND form_type = ?",
        (cik, form_type),
    ).fetchone()
    if row:
        return {"date": row[0], "accession": row[1]}
    return None


def upsert_filing(
    conn: sqlite3.Connection,
    cik: str,
    ticker: str,
    form_type: str,
    filing_date: str,
    accession: str,
) -> None:
    conn.execute(
        """INSERT INTO filing_tracker (cik, ticker, form_type, last_filing_date, accession_number, checked_at)
           VALUES (?, ?, ?, ?, ?, ?)
           ON CONFLICT(cik, form_type) DO UPDATE SET
               last_filing_date = excluded.last_filing_date,
               accession_number = excluded.accession_number,
               checked_at       = excluded.checked_at""",
        (cik, ticker, form_type, filing_date, accession, str(date.today())),
    )


# ---------------------------------------------------------------------------
# SEC EDGAR fetch
# ---------------------------------------------------------------------------

_session = requests.Session()
_session.headers.update({"User-Agent": EDGAR_UA, "Accept": "application/json"})


def fetch_recent_filings(cik: str) -> dict[str, dict] | None:
    """
    Return {form_type: {date, accession}} for the most recent 10-K and 10-Q.
    Returns None on fetch failure.
    """
    url = EDGAR_SUBMISSIONS.format(cik=cik)
    try:
        resp = _session.get(url, timeout=15)
        resp.raise_for_status()
    except requests.RequestException as e:
        print(f"  WARNING: EDGAR fetch failed for CIK {cik}: {e}")
        return None

    data = resp.json()
    recent = data.get("filings", {}).get("recent", {})
    forms = recent.get("form", [])
    dates = recent.get("filingDate", [])
    accessions = recent.get("accessionNumber", [])

    result: dict[str, dict] = {}
    for form, d, acc in zip(forms, dates, accessions):
        if form in FORMS and form not in result:
            result[form] = {"date": d, "accession": acc}
        if len(result) == len(FORMS):
            break

    return result


# ---------------------------------------------------------------------------
# Core check logic
# ---------------------------------------------------------------------------


def check_all(conn: sqlite3.Connection, init_mode: bool = False) -> list[dict]:
    """
    Check every stock.  Returns a list of new-filing dicts.
    In init_mode we seed the tracker without flagging anything as new.
    """
    stocks = get_tracked_stocks(conn)
    total = len(stocks)
    new_filings: list[dict] = []

    print(
        f"\nChecking {total} stocks against SEC EDGAR ({'seeding' if init_mode else 'monitoring'} mode) ...\n"
    )

    for idx, stock in enumerate(stocks, 1):
        ticker = stock["ticker"]
        cik = stock["cik"]
        name = stock["name"]

        print(f"\r  {idx}/{total}  {ticker:<8}", end="", flush=True)

        edgar_filings = fetch_recent_filings(cik)
        time.sleep(SLEEP_BETWEEN_REQUESTS)

        if edgar_filings is None:
            continue

        for form_type, filing in edgar_filings.items():
            known = get_known_filing(conn, cik, form_type)
            is_new = (known is None) or (filing["date"] > known["date"])

            if is_new and not init_mode:
                new_filings.append(
                    {
                        "ticker": ticker,
                        "name": name,
                        "cik": cik,
                        "form_type": form_type,
                        "date": filing["date"],
                        "accession": filing["accession"],
                        "prev_date": known["date"] if known else "—",
                        "url": EDGAR_FILING_URL.format(cik=cik, form=form_type),
                    }
                )

            # Always update tracker to latest known filing
            upsert_filing(
                conn, cik, ticker, form_type, filing["date"], filing["accession"]
            )

        conn.commit()

    print()  # newline after progress
    return new_filings


# ---------------------------------------------------------------------------
# Console report
# ---------------------------------------------------------------------------


def print_report(new_filings: list[dict]) -> None:
    if not new_filings:
        print("\nNo new 10-K or 10-Q filings detected.\n")
        return

    print(f"\n{'=' * 72}")
    print(f"  NEW FILINGS DETECTED: {len(new_filings)}")
    print(f"{'=' * 72}\n")

    # Group by form type
    for form in FORMS:
        group = [f for f in new_filings if f["form_type"] == form]
        if not group:
            continue
        print(f"  {form}  ({len(group)} new)\n")
        for f in sorted(group, key=lambda x: x["date"], reverse=True):
            print(
                f"    {f['ticker']:<8}  {f['name'][:40]:<40}  {f['date']}  (prev: {f['prev_date']})"
            )
            print(f"             {f['url']}")
        print()


# ---------------------------------------------------------------------------
# xlsx report
# ---------------------------------------------------------------------------


def write_xlsx(new_filings: list[dict]) -> str:
    today_str = date.today().strftime("%Y%m%d")
    path = os.path.join(OUTPUT_DIR, f"new_filings_{today_str}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "New Filings"

    THIN = Side(style="thin")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    DARK_FILL = PatternFill("solid", fgColor="1F4E79")
    GREEN_FILL = PatternFill("solid", fgColor="D4EDDA")

    # Title
    ws.cell(
        row=1,
        column=1,
        value=f"SEC EDGAR New Filings — {date.today().strftime('%B %d, %Y')}",
    ).font = Font(bold=True, size=13)
    ws.cell(
        row=2, column=1, value=f"{len(new_filings)} new 10-K/10-Q filings detected"
    ).font = Font(italic=True, color="666666")

    headers = [
        "Ticker",
        "Company",
        "CIK",
        "Form",
        "Filed",
        "Prior Filing",
        "EDGAR Link",
    ]
    HDR_ROW = 4
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=HDR_ROW, column=col, value=hdr)
        c.font = HEADER_FONT
        c.fill = DARK_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    for row_idx, f in enumerate(
        sorted(new_filings, key=lambda x: (x["form_type"], x["date"]), reverse=True),
        HDR_ROW + 1,
    ):
        row_vals = [
            f["ticker"],
            f["name"],
            f["cik"],
            f["form_type"],
            f["date"],
            f["prev_date"],
            f["url"],
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill = GREEN_FILL
            c.border = BORDER
            if col == 7:  # hyperlink the URL column
                c.hyperlink = val
                c.value = "View on EDGAR"
                c.font = Font(color="0563C1", underline="single")
            elif col <= 2:
                c.alignment = Alignment(horizontal="left")
            else:
                c.alignment = Alignment(horizontal="center")

    col_widths = [10, 36, 14, 8, 12, 12, 18]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    args = sys.argv[1:]
    init_mode = "--init" in args
    xlsx_mode = "--report" in args or "--xlsx" in args

    conn = open_db()

    try:
        new_filings = check_all(conn, init_mode=init_mode)
    finally:
        conn.close()

    if init_mode:
        print(f"Tracker seeded.  Run without --init to monitor for new filings.\n")
        return

    print_report(new_filings)

    if xlsx_mode and new_filings:
        path = write_xlsx(new_filings)
        print(f"xlsx report saved: {path}\n")
    elif xlsx_mode and not new_filings:
        print("No new filings — xlsx not written.\n")


if __name__ == "__main__":
    main()
