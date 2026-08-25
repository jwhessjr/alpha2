"""
Earnings whisper truth--
Price dances with hope and fear,
Worth hides in the mist.

S&P 500 / Russell 2000 batch valuation using FCFF DCF model.
Outputs results to value_<index>_YYYYMMDD.xlsx, sorted by margin of safety.


NOTE: Alpha Vantage free tier allows ~25 API requests/day (5/min).
      Each stock requires ~4-5 calls; use --limit N to cap the number of stocks.
      Usage: python av_fcff_2.py [--limit N] [--growth N]
"""

import argparse
from dataclasses import dataclass
from datetime import date
import sqlite3
import sys as _sys
from pathlib import Path as _Path
# This file is intentionally kept identical to
# /Users/jhess/Development/Alpha2/src/av_fcff_2.py (see docs/known_errors.md,
# 2026-07-14 — that copy is actively used by stock_analysis.py, not orphaned).
# hg_dcflib.py has its own separate, manually-synced copy in each location
# (per CLAUDE.md), so we APPEND ~/HessGrp/lib/ rather than inserting at the
# front — this makes it a fallback for logging_setup.py (which only exists
# in HessGrp/lib), without shadowing a same-directory hg_dcflib.py copy.
_sys.path.append(str(_Path.home() / "HessGrp" / "lib"))
import hg_dcflib
from config import INTRINIO_KEY
import json
import logging
import os
import sys
import time
import traceback
import io
import pandas as pd
import requests

from logging_setup import make_logger, LONG_FMT

if getattr(sys, "frozen", False):
    _log_dir = os.path.join(os.path.dirname(sys.executable), "data")
else:
    _log_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
os.makedirs(_log_dir, exist_ok=True)

logger = make_logger(__name__, os.path.join(_log_dir, "value.log"),
                     stream_level=logging.WARNING, fmt=LONG_FMT)


# ---------------------------------------------------------------------------
# Constants (fetched once at startup)
# ---------------------------------------------------------------------------

MY_API_KEY = os.environ.get("ALPHA_VANTAGE_API_KEY")
FRED_KEY = os.environ.get("FRED_API_KEY")
if not MY_API_KEY:
    raise EnvironmentError(
        "ALPHA_VANTAGE_API_KEY is not set. Run: export ALPHA_VANTAGE_API_KEY='your_key'"
    )
if not FRED_KEY:
    raise EnvironmentError(
        "FRED_API_KEY is not set. Run: export FRED_API_KEY='your_key'"
    )

MARGINAL_TAX_RATE = 0.26
GROWTH_PERIOD = 5  # high-growth years; override with --growth N

# AV->Intrinio migration Phase 3 (2026-08-24): "intrinio" is now the
# production default for the 5 fetch wrappers below, with automatic
# per-call fallback to AV if an Intrinio fetch fails for a given ticker
# (see _fetch_with_fallback()) -- not just an available alternative via
# --provider av. AV's own subscription runs regardless through April 2027,
# and Phase 2's real 3-way comparison against SEC EDGAR ground truth found
# each vendor fails on different tickers for different reasons, so combining
# both should reduce total nightly failures versus either alone. Full
# rationale: docs/decisions.md, "Data provider: Intrinio becomes primary".
DATA_PROVIDER = "intrinio"

# DB path — override via $VALUATION_DB env var or --db argument
DEFAULT_DB = os.environ.get("VALUATION_DB", "/Volumes/Financial_Data/valuation.db")

# Deferred to main() so a slow/failed network call at startup doesn't block
# argument parsing or prevent a scheduled job from reporting a clean error.
# Values below are used as fallbacks if the live fetch fails.
EQ_PREM: float = 0.0472    # Damodaran Jan 2026 US ERP fallback
RISK_FREE: float = 0.0425  # approximate 10-yr Treasury fallback
STABLE_GROWTH: float = 0.030  # long-run US nominal GDP growth rate (Damodaran ceiling: <= risk-free)
EQUITY_OVERRIDE: float | None = None  # set via --equity-override; bypasses AV balance sheet equity pull

# Moat-gated stable-phase ROIC blend (decided 2026-08-01, see docs/decisions.md
# and docs/known_errors.md). weight=0 (moat_rating "None" or no moat_scores
# row) reproduces pure g/WACC convergence — Damodaran's own conservative
# default, no persistent excess returns. weight=1 (Wide moat, sustained
# >= MOAT_CONFIDENCE_YEARS) assumes full ROIC persistence. Narrow/Questionable
# interpolate. Confidence in a moat rating is throttled by years_above_wacc
# (scripts/moat_score.py) so one good quarter can't buy a decade of credit.
MOAT_CONFIDENCE_YEARS = 5
MOAT_BASE_WEIGHT = {"Wide": 1.0, "Narrow": 0.5, "Questionable": 0.15, "None": 0.0}

# Capital-light-compounder ROIC gate (decided 2026-08-10, see docs/decisions.md
# and docs/known_errors.md). calc_return_on_capital()'s denominator (equity +
# debt - cash) goes negative for cash-rich, heavy-buyback companies (AZO,
# EXPE, etc.), producing a spurious sign-flipped ROIC. Rather than exclude the
# whole category, a company with positive earnings but non-positive invested
# capital must clear all three gates below to be treated as a wealth creator;
# failing any of them degrades to the same flagged-skip treatment as negative
# book equity. See calc_gated_return_on_capital().
WEALTH_GATE_MIN_YEARS = 3               # years of positive EBIT required
WEALTH_GATE_MIN_INTEREST_COVERAGE = 4.0  # EBIT / interest expense
WEALTH_GATE_MIN_YEARS_ABOVE_WACC = 3     # corroborating moat_scores track record

# ROIC corroboration flag (decided 2026-08-20, docs/decisions.md) — for the
# normal positive-invested-capital case, the current-period return_on_capital
# is used as-is with no check against the company's own longer-run history.
# Found live: NUTX (implied ROIC ~121% vs. moat_scores' own -15.3% 10-year
# average) and NRC (~151% vs. 47.7%) both hit the 30% growth-rate cap on a
# single strong quarter with no reference to their track record. An absolute
# spread, not a ratio, since NUTX's own average is negative (a ratio is
# undefined/meaningless there). Flags via `notes`, never corrects the
# computed number — see calc_gated_return_on_capital()'s docstring.
ROIC_CORROBORATION_MAX_SPREAD = 0.50    # percentage points above moat_scores' avg_roic
ROIC_CORROBORATION_MIN_DATA_YEARS = 3   # avg_roic needs enough history to be a credible baseline


# ---------------------------------------------------------------------------
# Data class
# ---------------------------------------------------------------------------


@dataclass
class Stock_Value:
    ticker: str
    valuation_date: str
    ent_name: str
    industry: str
    cik: str
    beta: float
    market_cap: float
    price: float
    shares_outstanding: float
    risk_free_rate: float
    eq_premium: float
    growth_rate: float
    cost_of_capital: float
    wealth_pc: float
    fcff_value: float
    terminal_value: float
    share_value: float
    margin_of_safety: float
    margin_of_safety_pc: float
    target_price: float
    earnings_yield: float = 0.0
    dividend_yield: float = 0.0
    notes: str = ""
    analyst_count: int = 0


# ---------------------------------------------------------------------------
# S&P 500 ticker list
# ---------------------------------------------------------------------------


def get_sp500_tickers() -> list:
    """
    Return the current S&P 500 constituent list.

    Reads from data/sp500_tickers.csv if available (produced by ticker_lists).
    Falls back to a live Wikipedia fetch if the file is not found.
    """
    csv_path = os.path.join(_log_dir, "sp500_tickers.csv")
    if os.path.exists(csv_path):
        df = pd.read_csv(csv_path)
        tickers = df["Ticker"].dropna().astype(str).str.strip().tolist()
        logger.info(f"Loaded {len(tickers)} S&P 500 tickers from {csv_path}")
        return tickers

    # ── Fallback: live fetch from Wikipedia ────────────────────────────────
    logger.warning("sp500_tickers.csv not found — fetching live from Wikipedia")
    url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
    headers = {"User-Agent": "Mozilla/5.0 (compatible; research-bot/1.0)"}
    resp = requests.get(url, headers=headers, timeout=15)
    resp.raise_for_status()
    tables = pd.read_html(io.StringIO(resp.text), header=0)
    df = tables[0]
    tickers = df["Symbol"].str.replace(".", "-", regex=False).tolist()
    logger.info(f"Fetched {len(tickers)} S&P 500 tickers from Wikipedia")
    return tickers


def get_russell2000_tickers() -> list:
    """
    Return the current Russell 2000 constituent list.

    Priority:
      1. data/russell2000_tickers.csv (produced by ticker_lists.py — most accurate)
      2. Live iShares IWM CSV (requires no-auth access — may be blocked)
      3. valuation.db full ticker universe (fallback when iShares is unavailable)
    """
    csv_path = os.path.join(_log_dir, "russell2000_tickers.csv")
    if os.path.exists(csv_path):
        df = pd.read_csv(csv_path)
        tickers = df["Ticker"].dropna().astype(str).str.strip().tolist()
        logger.info(f"Loaded {len(tickers)} Russell 2000 tickers from {csv_path}")
        return tickers

    # ── Fallback 1: live fetch from iShares ───────────────────────────────
    logger.warning("russell2000_tickers.csv not found — trying live iShares fetch")
    try:
        url = (
            "https://www.ishares.com/us/products/239710/ishares-russell-2000-etf/"
            "1467271812596.ajax?fileType=csv&fileName=IWM_holdings&dataType=fund"
        )
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": "https://www.ishares.com/",
        }
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        if resp.text.lstrip().startswith("<"):
            raise ValueError("iShares returned HTML instead of CSV — direct download is blocked")
        lines = resp.text.splitlines()
        header_idx = next(
            (i for i, line in enumerate(lines) if "Ticker" in line and "Name" in line),
            None,
        )
        if header_idx is None:
            raise ValueError("Could not locate Ticker header row in iShares CSV")
        df = pd.read_csv(io.StringIO("\n".join(lines[header_idx:])))
        tickers = (
            df["Ticker"]
            .dropna()
            .astype(str)
            .str.strip()
            .str.replace(".", "-", regex=False)
            .pipe(lambda s: s[s.str.match(r"^[A-Z]{1,5}(-[A-Z]+)?$")])
            .tolist()
        )
        logger.info(f"Fetched {len(tickers)} Russell 2000 tickers from iShares")
        return tickers
    except Exception as e:
        logger.warning(f"iShares live fetch failed ({e}) — falling back to valuation.db universe")

    # ── Fallback 2: use all tickers already in valuation.db ───────────────
    db_path = os.environ.get("VALUATION_DB", DEFAULT_DB)
    if not os.path.exists(db_path):
        raise RuntimeError(
            f"russell2000_tickers.csv missing, iShares blocked, and valuation.db not found at {db_path}. "
            "Run ticker_lists.py to generate the CSV file."
        )
    conn = sqlite3.connect(db_path)
    rows = conn.execute("SELECT DISTINCT ticker FROM valuation ORDER BY ticker").fetchall()
    conn.close()
    tickers = [r[0] for r in rows]
    logger.warning(
        f"Using {len(tickers)} tickers from valuation.db as Russell 2000 proxy. "
        "Run ticker_lists.py to refresh russell2000_tickers.csv for an accurate constituent list."
    )
    print(
        f"\n  NOTE: iShares CSV unavailable. Running against {len(tickers)} tickers already in "
        "valuation.db.\n  For a fresh Russell 2000 list, run: python3 ticker_lists.py\n"
    )
    return tickers


# ---------------------------------------------------------------------------
# Filings-driven ticker list
# ---------------------------------------------------------------------------


def get_tickers_from_filings(path: str) -> list[str]:
    """
    Return a de-duplicated, order-preserved list of tickers from a filings file.

    Supported formats (detected by extension):
      .xlsx  — legacy sec_monitor output (header on row 4, "Ticker" column)
      .json  — sec_daily_index output: {"tickers": ["AAPL", "JBL", ...]}
               or a plain JSON array: ["AAPL", "JBL", ...]
      .txt   — one ticker per line, blank lines and # comments ignored
    """
    import json as _json

    ext = os.path.splitext(path)[1].lower()

    if ext == ".xlsx":
        df = pd.read_excel(path, header=3)   # row 4 (0-indexed row 3) is the header
        tickers = df["Ticker"].dropna().str.strip().str.upper().unique().tolist()

    elif ext == ".json":
        with open(path, "r") as f:
            data = _json.load(f)
        if isinstance(data, dict):
            raw = data.get("tickers", [])
        elif isinstance(data, list):
            raw = data
        else:
            raise ValueError(f"Unrecognised JSON structure in {path}")
        tickers = list(dict.fromkeys(t.strip().upper() for t in raw if t.strip()))

    elif ext == ".txt":
        with open(path, "r") as f:
            lines = f.readlines()
        tickers = list(dict.fromkeys(
            line.strip().upper()
            for line in lines
            if line.strip() and not line.strip().startswith("#")
        ))

    else:
        raise ValueError(f"Unsupported filings file format: {ext}  (expected .xlsx, .json, or .txt)")

    logger.info(f"Loaded {len(tickers)} tickers from filings file: {path}")
    return tickers


# ---------------------------------------------------------------------------
# Financial statement helpers
# ---------------------------------------------------------------------------


def _fetch_with_fallback(ticker, intrinio_fn, av_fn, label):
    """
    Phase 3 (2026-08-24): when DATA_PROVIDER == "intrinio" (the production
    default), try Intrinio first and automatically fall back to AV if the
    Intrinio call fails for this ticker -- not just AV being reachable via
    --provider av. Phase 2's real 3-way comparison against SEC EDGAR found
    each vendor fails on different tickers for different reasons (Intrinio:
    genuine coverage gaps on foreign-domiciled names; AV: the cash-corruption/
    coverage pattern that motivated the migration), so combining both
    automatically should reduce total nightly failures versus either alone.
    See docs/decisions.md, "Data provider: Intrinio becomes primary".

    When DATA_PROVIDER == "av" explicitly (manual/shadow-mode AV-only runs),
    no fallback applies -- calls AV directly, unchanged from Phase 1/2.
    """
    if DATA_PROVIDER != "intrinio":
        return av_fn()
    try:
        return intrinio_fn()
    except Exception as exc:
        logger.warning(
            f"{ticker}: Intrinio {label} fetch failed ({exc}) — falling back to AV."
        )
        return av_fn()


def income_statement(ticker, api_key):
    return _fetch_with_fallback(
        ticker,
        lambda: hg_dcflib.get_inc_stmnt_intrinio(ticker, INTRINIO_KEY),
        lambda: hg_dcflib.get_inc_stmnt(ticker, api_key),
        "income statement",
    )


def balance_sheet(ticker, api_key, is_financial_or_reit: bool = False):
    return _fetch_with_fallback(
        ticker,
        lambda: hg_dcflib.get_bal_sheet_intrinio(ticker, INTRINIO_KEY, is_financial_or_reit=is_financial_or_reit),
        lambda: hg_dcflib.get_bal_sheet(ticker, api_key, is_financial_or_reit=is_financial_or_reit),
        "balance sheet",
    )


def cash_flow_statement(ticker, api_key):
    return _fetch_with_fallback(
        ticker,
        lambda: hg_dcflib.get_cash_flow_intrinio(ticker, INTRINIO_KEY),
        lambda: hg_dcflib.get_cash_flow(ticker, api_key),
        "cash flow statement",
    )


def research_and_development(ticker, rd_years, api_key):
    return _fetch_with_fallback(
        ticker,
        lambda: hg_dcflib.get_rAndD_intrinio(ticker, rd_years, INTRINIO_KEY),
        lambda: hg_dcflib.get_rAndD(ticker, rd_years, api_key),
        "R&D",
    )


# Populated by prefetch_quotes() before a batch run — see that function's
# docstring and docs/known_errors.md (2026-07-22) for why quote and
# fundamentals calls are deliberately kept out of the same time window.
# Keyed by (ticker, DATA_PROVIDER) so a shadow-mode comparison that runs both
# providers within one process (e.g. a Phase 2 diff script) can't serve a
# cached AV quote back for an Intrinio-provider call or vice versa. A
# fallback-to-AV quote (Phase 3) is cached under the "intrinio" provider key
# it was requested under, not "av" -- it's what enterprise_quote() actually
# returned for that (ticker, DATA_PROVIDER) combination.
_QUOTE_CACHE: dict = {}


def enterprise_quote(ticker, api_key):
    cache_key = (ticker, DATA_PROVIDER)
    if cache_key in _QUOTE_CACHE:
        return _QUOTE_CACHE[cache_key]
    return _fetch_with_fallback(
        ticker,
        lambda: hg_dcflib.get_quote_intrinio(ticker, INTRINIO_KEY),
        lambda: hg_dcflib.get_quote(ticker, api_key),
        "quote",
    )


def prefetch_quotes(tickers: list, api_key: str) -> None:
    """
    Fetch GLOBAL_QUOTE/OVERVIEW (via hg_dcflib.get_quote) for every ticker in
    one contiguous batch, before any fundamentals calls (INCOME_STATEMENT/
    BALANCE_SHEET/CASH_FLOW) begin, and cache the results in _QUOTE_CACHE.

    Alpha Vantage support confirmed (2026-07-22) that interleaving GLOBAL_QUOTE
    (real-time, entitlement-gated) with fundamentals calls for the same symbol
    within the same short window can trip a per-minute micro-throttle separate
    from the account's headline RPM cap — even when the overall request rate is
    far under that cap (our own logs showed ~11-12 req/min, well under the 75/
    min premium limit, still failing ~37% of tickers). Every one of our 6
    duplicated valuation paths calls the shared enterprise_quote() wrapper
    above, so caching there fixes all 6 without touching any of them — see
    value_bank_stock, _value_stock_fcff, value_reit_stock,
    _value_bank_stock_detail, _value_reit_stock_detail, _value_stock_detail_fcff.

    Per-ticker failures here are logged and simply left out of the cache —
    enterprise_quote() falls back to a live call for anything not cached, so a
    prefetch miss degrades to the old (interleaved) behavior for that one
    ticker rather than blocking the whole run.
    """
    total = len(tickers)
    bar_width = 40
    start_time = time.time()
    for idx, ticker in enumerate(tickers, 1):
        try:
            _QUOTE_CACHE[(ticker, DATA_PROVIDER)] = _fetch_with_fallback(
                ticker,
                lambda t=ticker: hg_dcflib.get_quote_intrinio(t, INTRINIO_KEY),
                lambda t=ticker: hg_dcflib.get_quote(t, api_key),
                "quote",
            )
        except Exception as e:
            # Both Intrinio and its AV fallback failed (or DATA_PROVIDER=="av"
            # and AV itself failed) -- logged and left out of the cache;
            # enterprise_quote()'s own fallback-aware live call covers this
            # ticker later, same degrade-gracefully behavior as before.
            logger.warning(f"Prefetch quote failed for {ticker}: {e}")
        filled = int(bar_width * idx / total)
        bar = "#" * filled + "-" * (bar_width - filled)
        elapsed = int(time.time() - start_time)
        h, rem = divmod(elapsed, 3600)
        m, s = divmod(rem, 60)
        print(f"\r  quotes {idx}/{total} [{bar}] {h:02d}:{m:02d}:{s:02d}", end="", flush=True)
    print()


def get_excluded_tickers() -> set:
    """
    Load the permanently-excluded ticker set from data/excluded_tickers.json.

    Found 2026-07-23: this file was previously read only by Iggy's SKILL.md
    orchestration (hess_group/scheduled/iggy-valuation-update/SKILL.md), and
    only to strip excluded tickers from the *next day's* retry file — never to
    skip them within the same run. Every batch run was still attempting (and,
    since the 2026-07-22 second-pass retry queue, re-attempting) every
    excluded ticker before that filtering ever kicked in. This wires the same
    file directly into the batch loop so excluded tickers are skipped before
    the first attempt, not just before tomorrow's retry.

    Resolves the file the same way hg_dcflib.py resolves reference_data/: try
    the path relative to this script's own location first (so a duplicated
    copy in Development/Alpha2 would be preferred there), falling back to the
    fixed ~/HessGrp/data/ path since this file is HessGrp-specific and has no
    Alpha2 counterpart today. See docs/known_errors.md 2026-07-14 entry for
    the same fallback pattern used for logging_setup.py/hg_dcflib.py.
    """
    candidates = [
        _Path(os.path.abspath(__file__)).parent.parent / "data" / "excluded_tickers.json",
        _Path.home() / "HessGrp" / "data" / "excluded_tickers.json",
    ]
    for path in candidates:
        if path.exists():
            try:
                with open(path) as f:
                    return set(t.upper() for t in json.load(f).get("tickers", []))
            except Exception as e:
                logger.warning(f"Could not load excluded_tickers.json at {path}: {e}")
    return set()


def get_moat_weight(ticker: str, db_path: str | None = None) -> float:
    """
    Blend weight toward full stable-phase ROIC persistence, gated on
    scripts/moat_score.py's moat_rating and years_above_wacc — see
    MOAT_BASE_WEIGHT/MOAT_CONFIDENCE_YEARS above and docs/decisions.md
    "Moat-gated stable-phase ROIC assumption" (decided 2026-08-01).

    Missing moat_scores row, missing table, or any read failure all degrade
    to weight 0.0 — i.e. pure WACC-convergence, the same behavior as before
    this feature existed. A missing/stale moat score should never make a
    valuation *more* optimistic than the conservative default.
    """
    try:
        conn = sqlite3.connect(db_path or DEFAULT_DB, timeout=10)
        row = conn.execute(
            "SELECT moat_rating, years_above_wacc FROM moat_scores WHERE ticker = ?",
            (ticker,),
        ).fetchone()
        conn.close()
    except Exception as e:
        logger.debug(f"{ticker}: could not read moat_scores ({e}) — moat weight 0.0")
        return 0.0

    if row is None:
        return 0.0

    moat_rating, years_above_wacc = row
    base = MOAT_BASE_WEIGHT.get(moat_rating, 0.0)
    confidence = min(years_above_wacc / MOAT_CONFIDENCE_YEARS, 1.0) if years_above_wacc else 0.0
    return base * confidence


def get_moat_corroboration(
    ticker: str, db_path: str | None = None
) -> tuple[str | None, int | None, float | None, int | None]:
    """
    Fetch (moat_rating, years_above_wacc, avg_roic, data_years) from
    moat_scores — the raw fields behind get_moat_weight(), used by
    calc_gated_return_on_capital()'s Gate 3 (see docs/decisions.md
    "Capital-light compounder ROIC gate") and its ROIC-corroboration flag
    (docs/decisions.md, decided 2026-08-20). moat_score.py's own avg_roic
    already skips non-positive-invested-capital years
    (scripts/moat_score.py: score_roic()), so it's reused directly rather
    than inventing a second capital-efficiency heuristic here.

    Same failure contract as get_moat_weight(): missing row, missing table,
    or any read failure all degrade to (None, None, None, None) — never more
    optimistic than "gate fails"/"no corroboration" on a data gap.
    """
    try:
        conn = sqlite3.connect(db_path or DEFAULT_DB, timeout=10)
        row = conn.execute(
            "SELECT moat_rating, years_above_wacc, avg_roic, data_years "
            "FROM moat_scores WHERE ticker = ?",
            (ticker,),
        ).fetchone()
        conn.close()
    except Exception as e:
        logger.debug(f"{ticker}: could not read moat_scores ({e}) — no gate corroboration")
        return None, None, None, None

    if row is None:
        return None, None, None, None
    return row[0], row[1], row[2], row[3]


# ---------------------------------------------------------------------------
# Industry classification helpers
# ---------------------------------------------------------------------------

_FINANCIAL_KEYWORDS = {
    "bank",
    "banks",
    "financial services",
    "insurance",
    "brokerage",
    "investment banking",
    "thrift",
    "savings",
    "credit",
    "mortgage",
    "asset management",
}

_INSURANCE_KEYWORDS = {"insurance", "reinsurance", "surety", "title insurance"}

_REIT_KEYWORDS = {"reit", "real estate investment trust"}
_REIT_INDUSTRY_PREFIXES = ("retail (reit", "r.e.i.t.")

# Growth rate floors by REIT sub-type (AV industry string keyword → rate).
# Reflects contractual lease escalators and structural growth independent of
# retained earnings. None = mortgage REIT; flag for manual review.
_REIT_SUBTYPE_GROWTH: dict[str, float | None] = {
    "specialty":   0.020,  # broad bucket (data centers, self-storage, etc.) — conservative default
    "industrial":  0.010,
    "residential": 0.010,
    "healthcare":  0.015,
    "diversified": 0.005,
    "hotel":       0.000,
    "retail":      0.000,  # conservative; net lease overrides below
    "office":      0.000,
    "mortgage":    None,   # not property income — AFFO DDM not applicable
}

# Ticker-level overrides for known sub-types where the industry keyword is too coarse.
# Tower REITs and data centers: CPI escalators + colocation growth → 3%.
# Net lease REITs: contractual annual escalators (1–2%) → 1.5%.
_REIT_TICKER_GROWTH_OVERRIDE: dict[str, float] = {
    "SBAC": 0.030, "AMT": 0.030, "CCI": 0.030,          # tower
    "EQIX": 0.030, "DLR": 0.030, "CONE": 0.030,          # data center
    "O":    0.015, "NNN": 0.015, "STOR": 0.015,           # net lease
    "ELS":  0.020, "SUI": 0.020,                           # manufactured housing
    "WY":   0.015, "PCH": 0.015,                           # timber (biological growth proxy)
}


def reit_subtype_growth(ticker: str, industry: str) -> float | None:
    """Return DDM growth rate floor for a REIT based on sub-type.

    Returns None for mortgage REITs (AFFO DDM not applicable — flag for manual review).
    Ticker-level override takes precedence over industry keyword match.
    """
    if ticker in _REIT_TICKER_GROWTH_OVERRIDE:
        return _REIT_TICKER_GROWTH_OVERRIDE[ticker]
    low = industry.lower()
    for keyword, rate in _REIT_SUBTYPE_GROWTH.items():
        if keyword in low:
            return rate
    return 0.010  # unknown sub-type — conservative fallback


def is_financial_firm(industry: str) -> bool:
    """Return True if the industry is a financial firm requiring FCFE valuation."""
    low = industry.lower()
    return any(kw in low for kw in _FINANCIAL_KEYWORDS)


def is_insurance_firm(industry: str) -> bool:
    """Return True for insurance companies that need normalized NI."""
    low = industry.lower()
    return any(kw in low for kw in _INSURANCE_KEYWORDS)


def is_reit(industry: str) -> bool:
    """Return True for REITs (pass-through entities requiring AFFO DDM).
    Excludes real estate services/development/brokerage firms — those use FCFF."""
    low = industry.lower()
    return (
        any(kw in low for kw in _REIT_KEYWORDS)
        or any(low.startswith(p) for p in _REIT_INDUSTRY_PREFIXES)
    )


def _normalized_net_income(net_income_list: list) -> tuple[float, int]:
    """
    Return (normalized_NI, years_used).
    Uses up to 5 years of NI, excluding any years where NI < 0
    (catastrophe or reserve-charge years) if at least 2 positive years exist.
    Falls back to simple average if most years are negative.
    """
    positives = [ni for ni in net_income_list if ni > 0]
    if len(positives) >= 2:
        avg = sum(positives) / len(positives)
        return avg, len(positives)
    # If fewer than 2 positive years, use all available years
    avg = sum(net_income_list) / len(net_income_list)
    return avg, len(net_income_list)


# ---------------------------------------------------------------------------
# Calculation functions
# ---------------------------------------------------------------------------


def calc_stable_beta(unlevered_beta):
    if unlevered_beta < 0.5:
        stable_beta = 0.8
    elif unlevered_beta > 1.5:
        stable_beta = 1.2
    else:
        stable_beta = 1.0
    logger.info(f"Stable beta = {stable_beta:,.3f}")
    return stable_beta


def calc_capital_expenditures(cash_flw):
    capex_years = cash_flw["capex"][:5]
    return sum(capex_years) / len(capex_years)


def calc_chng_wc(bal_sht):
    if len(bal_sht["total_current_assets"]) < 2:
        raise ValueError("Insufficient balance sheet history (need 2 years) to compute working capital change")
    curr_yr_nc_wc = (
        bal_sht["total_current_assets"][0] - bal_sht["cash_and_equivalents"][0]
    ) - (bal_sht["total_current_liabilities"][0] - bal_sht["short_term_debt"][0])
    prior_yr_nc_wc = (
        bal_sht["total_current_assets"][1] - bal_sht["cash_and_equivalents"][1]
    ) - (bal_sht["total_current_liabilities"][1] - bal_sht["short_term_debt"][1])
    return curr_yr_nc_wc - prior_yr_nc_wc


def capitalizerAndD(ticker, rd_years, api_key):
    rd_years = int(rd_years)
    if rd_years <= 1:
        # No R&D amortization for this industry — skip API call and return zeroed schedule
        return {
            "rAndDExpense": [0.0],
            "unamortized_percent": [0.0],
            "unamort_amount": [0.0],
            "RD_Asset_Value": 0.0,
            "Current_Year_Amortization": 0.0,
        }

    rdTable = research_and_development(ticker, rd_years, api_key)
    rd_dict, years_to_process = rdTable
    logger.info(f"rdTable = {rdTable}")
    rd_table = {}
    rd_expense = []
    unamort_percent = []
    unamort_amt = []
    amort_percentage = 1.0 / (rd_years - 1)

    current_year_total_amortization = 0
    for year in range(1, min(years_to_process, rd_years)):
        current_year_total_amortization += (
            rd_dict["research_and_development"][year] * amort_percentage
        )

    rd_asset_value = 0
    for year in range(years_to_process):
        expense = rd_dict["research_and_development"][year]
        percent_unamort = 1.0 - (amort_percentage * year)
        unamort = expense * percent_unamort
        rd_expense.append(expense)
        unamort_percent.append(percent_unamort)
        unamort_amt.append(unamort)
        rd_asset_value += unamort

    rd_table["rAndDExpense"] = rd_expense
    rd_table["unamortized_percent"] = unamort_percent
    rd_table["unamort_amount"] = unamort_amt
    rd_table["RD_Asset_Value"] = rd_asset_value
    rd_table["Current_Year_Amortization"] = current_year_total_amortization
    return rd_table


def calc_fcff(inc_stmnt, bal_sht, cash_flw, eff_tax_rate):
    ebiat = inc_stmnt["ebit"][0] * (1 - eff_tax_rate)
    logger.info(f"ebiat {ebiat:,.2f}")
    capex = calc_capital_expenditures(cash_flw)
    logger.info(f"Capex {capex:,.2f}")
    chng_nc_wc = calc_chng_wc(bal_sht)
    logger.info(f"Change WC {chng_nc_wc:,.2f}")
    depreciation = cash_flw["depreciation"][0]
    logger.info(f"Depreciation {depreciation:,.2f}")
    fcff = ebiat - capex + depreciation - chng_nc_wc
    logger.info(f"FCFF {fcff:,.2f}")
    return [ebiat, capex, chng_nc_wc, depreciation, fcff]


def calc_reinvestment(capex, depreciation, chng_nc_wc, amort_schedule):
    firm_reinvestment = (
        capex
        - depreciation
        + chng_nc_wc
        + amort_schedule["rAndDExpense"][0]
        - amort_schedule["Current_Year_Amortization"]
    )
    logger.info(f"Firm Reinvestment {firm_reinvestment:,.2f}")
    return firm_reinvestment


def calc_adj_ebiat(ebiat, amort_schedule):
    adjusted_ebiat = (
        ebiat
        + amort_schedule["rAndDExpense"][0]
        - amort_schedule["Current_Year_Amortization"]
    )
    logger.info(f"Adjusted ebiat {adjusted_ebiat:,.2f}")
    return adjusted_ebiat


def calc_adj_bv_equity(bal_sht, amort_schedule):
    if EQUITY_OVERRIDE is not None:
        base_equity = EQUITY_OVERRIDE
        logger.info(f"equity override active: using {base_equity:,.0f} instead of AV balance sheet")
    else:
        base_equity = bal_sht["total_stockholders_equity"][0]
    adjusted_bv_equity = base_equity + amort_schedule["RD_Asset_Value"]
    logger.info(f"adjusted BV Equity = {adjusted_bv_equity:,.2f}")
    return adjusted_bv_equity


def calc_bv_debt(bal_sht):
    bv_debt = bal_sht["short_term_debt"][0] + bal_sht["long_term_debt"][0]
    logger.info(f"BV Debt = {bv_debt:,.2f}")
    return bv_debt


def calc_tax_rate(inc_stmnt):
    income_before_tax = inc_stmnt["incomeBeforeTax"][0]
    if income_before_tax <= 0:
        # Loss-making company: effective rate is meaningless; use marginal rate
        logger.info(
            f"Negative/zero pre-tax income — using marginal tax rate {MARGINAL_TAX_RATE:.4f}"
        )
        return MARGINAL_TAX_RATE
    eff_tax_rate = inc_stmnt["income_tax_expense"][0] / income_before_tax
    # Clamp to [0, marginal rate] to prevent sign-flip in FCFF projections
    eff_tax_rate = min(max(eff_tax_rate, 0.0), MARGINAL_TAX_RATE)
    logger.info(f"Effective Tax Rate = {eff_tax_rate:,.4f}")
    return eff_tax_rate


def calc_return_on_capital(adjusted_ebiat, adjusted_bv_equity, bv_debt, bal_sht):
    return_on_capital = adjusted_ebiat / (
        adjusted_bv_equity + bv_debt - bal_sht["cash_and_equivalents"][0]
    )
    logger.info(f"ROIC = {return_on_capital:,.4f}")
    return return_on_capital


def calc_gated_return_on_capital(
    ticker: str,
    adjusted_ebiat: float,
    adjusted_bv_equity: float,
    bv_debt: float,
    bal_sht: dict,
    inc_stmnt: dict,
    db_path: str | None = None,
) -> tuple[float | None, str]:
    """
    Guarded wrapper around calc_return_on_capital() — see docs/decisions.md
    "Capital-light compounder ROIC gate" (decided 2026-08-10) and
    docs/known_errors.md for the full writeup.

    calc_return_on_capital()'s denominator (equity + debt - cash) goes
    negative for cash-rich, heavy-buyback companies (confirmed live: AZO,
    EXPE, CCSI, PRDO, INOD, and others), producing a spurious, sign-flipped
    ROIC that misclassifies genuinely profitable businesses as wealth
    destroyers and collapses the moat-gated terminal value to zero.

    Returns (return_on_capital, notes):
    - invested_capital > 0: passthrough to calc_return_on_capital(), with a
      new corroboration check (decided 2026-08-20, docs/decisions.md) — if
      the result exceeds moat_scores' own avg_roic by more than
      ROIC_CORROBORATION_MAX_SPREAD (and avg_roic has at least
      ROIC_CORROBORATION_MIN_DATA_YEARS of history to be a credible
      baseline), notes carries a flag but the computed return_on_capital is
      returned UNCHANGED — this is an annotation, not a correction, same
      "flag, don't silently exclude/correct" pattern as sector balance and
      the staleness warning (docs/decisions.md). Missing/thin moat data
      never triggers the flag — consistent with every other gate in this
      function degrading to "no signal" rather than "more optimistic."
    - adjusted_ebiat <= 0 (regardless of invested capital): (None, reason) —
      a real operating loss is a real signal; never overridden by this gate.
    - adjusted_ebiat > 0 and invested_capital <= 0 (the ambiguous case): must
      clear all three gates to be treated as a wealth creator —
        1. Durability: positive EBIT in each of the last WEALTH_GATE_MIN_YEARS
           years (inc_stmnt["ebit"] already carries up to 5 years, no new
           fetch).
        2. Interest coverage >= WEALTH_GATE_MIN_INTEREST_COVERAGE (mirrors
           the existing int_cover pattern used for get_default_spread()).
        3. Corroboration from moat_score.py's own ROIC series, which already
           skips non-positive-invested-capital years: a Wide/Narrow moat
           rating with >= WEALTH_GATE_MIN_YEARS_ABOVE_WACC years above WACC.
      All three pass: returns moat_score.py's avg_roic (reusing its
      already-guarded computation rather than inventing a second heuristic).
      Any gate fails: (None, reason naming the failed gate(s)).

    Callers must treat a None result by writing a flagged, zeroed Stock_Value
    with notes=<reason> instead of continuing the DCF -- this is now the ONLY
    negative-book-equity-adjacent skip path (see docs/known_errors.md
    2026-08-25 in HessGrp: a separate, cruder `adjusted_bv_equity < 0` guard
    used to fire before this function ever ran, unconditionally killing the
    DCF for any negative-book-equity ticker regardless of invested capital --
    removed so every such ticker gets a real shot at the three-gate test
    above instead of being assumed to fail it). docs/decisions.md's existing
    "downstream consumers filter on non-empty notes" rule already makes a
    None-result row correctly invisible to replacer.py's candidate queries
    and skips portfolio_monitor.py's elimination check — no consumer-side
    changes needed.
    """
    cash = bal_sht["cash_and_equivalents"][0]
    invested_capital = adjusted_bv_equity + bv_debt - cash

    if invested_capital > 0:
        roc = calc_return_on_capital(adjusted_ebiat, adjusted_bv_equity, bv_debt, bal_sht)
        _, _, moat_avg_roic, data_years = get_moat_corroboration(ticker, db_path)
        if (
            moat_avg_roic is not None
            and data_years is not None
            and data_years >= ROIC_CORROBORATION_MIN_DATA_YEARS
            and (roc - moat_avg_roic) > ROIC_CORROBORATION_MAX_SPREAD
        ):
            return roc, (
                f"Current ROIC ({roc:.1%}) exceeds moat_scores' "
                f"{data_years}yr average ({moat_avg_roic:.1%}) by more than "
                f"{ROIC_CORROBORATION_MAX_SPREAD:.0%} -- growth rate and "
                "wealth_pc are computed from the uncorroborated current "
                "figure; verify before trusting this as sustainable."
            )
        return roc, ""

    if adjusted_ebiat <= 0:
        return None, (
            "ROIC undefined -- negative/zero invested capital "
            f"({invested_capital:,.0f}) and non-positive earnings "
            f"(EBIAT {adjusted_ebiat:,.0f})"
        )

    # Ambiguous case: positive earnings, non-positive invested capital.
    ebit_hist = inc_stmnt.get("ebit", [])
    durable = (
        len(ebit_hist) >= WEALTH_GATE_MIN_YEARS
        and all(e > 0 for e in ebit_hist[:WEALTH_GATE_MIN_YEARS])
    )

    try:
        coverage = inc_stmnt["ebit"][0] / inc_stmnt["interest_expense"][0]
    except ZeroDivisionError:
        coverage = float("inf")  # no debt burden -- can't fail a coverage test
    covered = coverage >= WEALTH_GATE_MIN_INTEREST_COVERAGE

    moat_rating, years_above_wacc, moat_avg_roic, _ = get_moat_corroboration(ticker, db_path)
    corroborated = (
        moat_rating in ("Wide", "Narrow")
        and years_above_wacc is not None
        and years_above_wacc >= WEALTH_GATE_MIN_YEARS_ABOVE_WACC
        and moat_avg_roic is not None
    )

    if durable and covered and corroborated:
        logger.info(
            f"{ticker}: capital-light compounder -- standard ROIC undefined "
            f"(invested capital {invested_capital:,.0f}), cleared via "
            f"{WEALTH_GATE_MIN_YEARS}yr EBIT durability, {coverage:.1f}x "
            f"interest coverage, and moat_score.py's {moat_rating} rating "
            f"({years_above_wacc}yr above WACC) -- using moat_score's own "
            f"avg_roic ({moat_avg_roic:.1%}) in place of the undefined ratio"
        )
        return moat_avg_roic, ""

    reasons = []
    if not durable:
        reasons.append(f"< {WEALTH_GATE_MIN_YEARS}yr positive EBIT history")
    if not covered:
        reasons.append(f"interest coverage {coverage:.1f}x < {WEALTH_GATE_MIN_INTEREST_COVERAGE}x")
    if not corroborated:
        reasons.append("no corroborating Wide/Narrow moat rating with sufficient track record")
    return None, (
        f"ROIC undefined -- negative invested capital ({invested_capital:,.0f}), "
        "failed gate: " + "; ".join(reasons)
    )


def calc_growth_rate(reinvestment_rate, return_on_capital):
    growth_rate = reinvestment_rate * return_on_capital
    logger.info(f"Growth Rate = {growth_rate:,.4f}")
    return growth_rate


def _bank_growth_rate(roe: float, retention_ratio: float) -> float:
    """Shared by value_bank_stock()/_value_bank_stock_detail() — identical
    formula in both, extracted per the av_fcff_2.py consolidation plan."""
    return min(roe * retention_ratio, 0.30)


def _reit_growth_rate(roe: float, retention_ratio: float, subtype_floor: float):
    """Shared by value_reit_stock()/_value_reit_stock_detail() — identical
    formula in both, extracted per the av_fcff_2.py consolidation plan.

    Returns (growth_rate, retained_growth) — retained_growth is its own
    Excel report field in the detail path, not just an intermediate.
    subtype_floor itself is NOT computed here: batch resolves it before any
    data fetch (to skip mortgage REITs early), detail resolves it after —
    each call site keeps its own gating exactly where it already is.
    """
    retained_growth = min(roe * retention_ratio, 0.15)
    return max(retained_growth, subtype_floor), retained_growth


def calc_levered_beta(unlevered_beta, bv_debt, market_cap_equity, tax_rate, de_cap=None):
    """
    de_cap (optional): caps the D/E ratio used for re-levering at this value if
    the company's own D/E exceeds it — never raises D/E if the company's own is
    already lower. Used only for the stable/terminal-phase beta (pass the
    industry-average D/E via hg_dcflib.get_industry_de()) so an over-levered
    company's assumed perpetual leverage is capped toward a typical level,
    while an already-conservative company's IV is never inflated by assuming
    more leverage than it actually carries. See docs/decisions.md "Stable-phase
    capital structure" (decided 2026-07-22) — never pass de_cap for the
    explicit-period beta, which must always use the company's actual own D/E.
    """
    de_ratio = bv_debt / market_cap_equity if market_cap_equity > 0 else 0.0
    if de_cap is not None:
        de_ratio = min(de_ratio, de_cap)
    levered_beta = unlevered_beta * (1 + (1 - tax_rate) * de_ratio)
    logger.info(f"Levered beta = {levered_beta:,.4f} (unlevered {unlevered_beta:,.4f}, D/E {de_ratio:,.4f})")
    return levered_beta


def calc_discount_rate(inc_stmnt, bv_debt, market_cap_equity, beta, risk_free, eq_prem, de_cap=None):
    # Re-lever the industry (unlevered) beta to this company's own capital
    # structure before computing cost of equity — see docs/known_errors.md
    # (2026-07-14: previously used the raw unlevered beta directly in CAPM).
    # de_cap: pass hg_dcflib.get_industry_de(industry) when computing the
    # stable/terminal-phase rate (beta = stable_beta) only — never for the
    # explicit-period call. See docs/decisions.md "Stable-phase capital
    # structure" (decided 2026-07-22).
    levered_beta = calc_levered_beta(beta, bv_debt, market_cap_equity, MARGINAL_TAX_RATE, de_cap=de_cap)
    cost_of_equity = risk_free + (levered_beta * eq_prem)
    logger.info(f"COE = {cost_of_equity:,.4f}")

    try:
        int_cover = inc_stmnt["ebit"][0] / inc_stmnt["interest_expense"][0]
    except ZeroDivisionError:
        int_cover = 25

    logger.info(f"Interest Coverage = {int_cover}")
    def_spread = hg_dcflib.get_default_spread(int_cover)
    logger.info(f"Default Spread = {def_spread}")

    cost_of_debt = (risk_free + def_spread) * (1 - MARGINAL_TAX_RATE)
    logger.info(f"Cost of Debt = {cost_of_debt}")
    total_capital = market_cap_equity + bv_debt
    percent_debt = bv_debt / total_capital if total_capital > 0 else 0.5
    percent_equity = 1 - percent_debt

    cost_of_capital = (cost_of_debt * percent_debt) + (cost_of_equity * percent_equity)
    logger.info(f"Cost of Capital = {cost_of_capital:,.4f}")
    return cost_of_capital


def calc_expected_fcff(
    adjusted_ebit, eff_tax_rate, growth_rate, reinvestment_rate, growth_period
):
    # Grow EBIT each year, then derive EBIAT and FCFF.
    # Applying growth to EBIT (not EBIAT or FCFF) avoids sign errors when
    # the current FCFF is negative due to high reinvestment.
    ebit_n = []
    fcff_n = []
    for year in range(growth_period):
        if year == 0:
            ebit_n.append(adjusted_ebit * (1 + growth_rate))
        else:
            ebit_n.append(ebit_n[year - 1] * (1 + growth_rate))
        ebiat = ebit_n[year] * (1 - eff_tax_rate)
        fcff_n.append(ebiat * (1 - reinvestment_rate))
        logger.info(f"Expected FCFF year {year + 1} = {fcff_n[year]:,.2f}")
    return fcff_n


def calc_fcff_value(fcff_table, discount_rate, growth_period):
    fcff_value = 0
    for year in range(growth_period):
        fcff_pv = fcff_table[year] / ((1 + discount_rate) ** (year + 1))
        fcff_value += fcff_pv
    logger.info(f"FCFF Value = {fcff_value:,.2f}")
    return fcff_value


def calc_stable_reinvestment_rate(stable_growth, stable_cost_of_capital):
    """
    Stable-phase reinvestment rate: stable_growth / stable_cost_of_capital,
    i.e. ROIC converges to WACC in stable growth (no permanent excess
    returns) — the same assumption already used in value_bank_stock()'s and
    value_reit_stock()'s terminal value. See docs/known_errors.md 2026-07-31.
    """
    if stable_cost_of_capital <= 0:
        return 0.0
    return min(max(stable_growth / stable_cost_of_capital, 0.0), 1.0)


def calc_terminal_value(
    ebit_last, eff_tax_rate, stable_cost_of_capital, growth_cost_of_capital,
    stable_growth, growth_period, moat_weight=0.0, explicit_roic=None
):
    """
    Terminal value at the end of the explicit high-growth period, discounted
    back to present.

    Terminal-year FCFF is recomputed at the stable-phase reinvestment rate
    (see calc_stable_reinvestment_rate()) rather than carrying forward the
    explicit period's — typically much higher — reinvestment rate. The
    previous version grew the last explicit year's FCFF by stable_growth
    directly, with no adjustment to reinvestment even though growth had just
    dropped from (say) double digits to 3% — understating terminal FCFF (and
    hence terminal value, usually 70-90%+ of total DCF value) for any company
    whose explicit reinvestment rate exceeds what stable growth actually
    requires, which is true for nearly every profitable growth company. See
    docs/known_errors.md 2026-07-31.

    moat_weight / explicit_roic (optional): blend the stable-phase
    reinvestment rate toward full ROIC persistence (weight=1) instead of
    pure WACC-convergence (weight=0, the default — identical to the
    2026-07-31 behavior). See get_moat_weight() and docs/known_errors.md
    2026-08-01 "Moat-gated stable-phase ROIC assumption".
    """
    if moat_weight and explicit_roic is not None:
        assumed_stable_roic = stable_cost_of_capital + moat_weight * (explicit_roic - stable_cost_of_capital)
        if assumed_stable_roic > 0:
            stable_reinv_rate = min(max(stable_growth / assumed_stable_roic, 0.0), 1.0)
        else:
            stable_reinv_rate = 1.0
    else:
        stable_reinv_rate = calc_stable_reinvestment_rate(stable_growth, stable_cost_of_capital)
    terminal_ebit = ebit_last * (1 + stable_growth)
    terminal_ebiat = terminal_ebit * (1 - eff_tax_rate)
    fcff_terminal = terminal_ebiat * (1 - stable_reinv_rate)
    # Gordon Growth requires cost of capital > growth rate — otherwise this
    # denominator is zero or negative and terminal value is undefined. Not a
    # theoretical concern: a real, unguarded ZeroDivisionError hit CLMB, OSW,
    # RCKY during the 2026-07-31 Russell 2000 triage. See docs/known_errors.md.
    if stable_cost_of_capital <= stable_growth:
        raise ValueError(
            f"Stable-phase cost of capital ({stable_cost_of_capital:.4f}) is at or "
            f"below the stable growth rate ({stable_growth:.4f}) — terminal value "
            f"is undefined (requires cost of capital > growth)."
        )
    terminal_value = fcff_terminal / (stable_cost_of_capital - stable_growth)
    terminal_value_pv = terminal_value / ((1 + growth_cost_of_capital) ** growth_period)
    logger.info(f"Stable reinvestment rate = {stable_reinv_rate:,.4f}")
    logger.info(f"Terminal Value = {terminal_value_pv:,.2f}")
    return terminal_value_pv


def calc_intrinsic_value(
    fcff_pv, terminal_value_pv, cash_and_equivalents, bv_debt, shares_outstanding
):
    enterprise_value = fcff_pv + terminal_value_pv + cash_and_equivalents - bv_debt
    intrinsic_value = enterprise_value / shares_outstanding
    logger.info(f"Enterprise Value = {enterprise_value:,.2f}")
    logger.info(f"Intrinsic Value = {intrinsic_value:,.2f}")
    return intrinsic_value


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------


def create_table(conn):
    schema_sql = """CREATE TABLE IF NOT EXISTS valuation (
              ticker TEXT NOT NULL,
              valuation_date TEXT NOT NULL,
              ent_name TEXT NOT NULL,
              industry TEXT NOT NULL,
              cik TEXT NOT NULL DEFAULT '',
              beta REAL NOT NULL,
              market_cap REAL NOT NULL,
              price REAL NOT NULL,
              shares_outstanding REAL NOT NULL,
              risk_free_rate REAL NOT NULL,
              eq_premium REAL NOT NULL,
              growth_rate REAL NOT NULL,
              cost_of_capital REAL NOT NULL,
              wealth_pc REAL NOT NULL,
              fcff_value REAL NOT NULL,
              terminal_value REAL NOT NULL,
              share_value REAL NOT NULL,
              margin_of_safety REAL NOT NULL,
              margin_of_safety_pc REAL NOT NULL,
              target_price REAL NOT NULL DEFAULT 0,
              earnings_yield REAL NOT NULL DEFAULT 0,
              dividend_yield REAL NOT NULL DEFAULT 0,
              notes TEXT NOT NULL DEFAULT '',
              analyst_count INTEGER NOT NULL DEFAULT 0,
              PRIMARY KEY (ticker)
              );"""
    try:
        # Check if table exists with old (ticker, valuation_date) composite PK
        # and migrate if needed so that each ticker has only one row.
        row = conn.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='valuation'"
        ).fetchone()
        if row and "PRIMARY KEY (ticker, valuation_date)" in row[0]:
            logger.info("Migrating valuation table to single-ticker primary key ...")
            # DDL (ALTER/CREATE/DROP) bypasses Python's automatic transaction
            # management, so we disable it temporarily and use explicit SQL
            # BEGIN/COMMIT/ROLLBACK to guarantee atomicity.
            orig_isolation = conn.isolation_level
            conn.isolation_level = None
            try:
                conn.execute("BEGIN EXCLUSIVE")
                conn.execute("ALTER TABLE valuation RENAME TO valuation_old")
                conn.execute(schema_sql)
                # Explicit destination column list — cik and target_price are
                # new columns not present in the old table; they receive their
                # DEFAULT values ('', 0) automatically.
                conn.execute("""
                    INSERT INTO valuation (
                        ticker, valuation_date, ent_name, industry, beta, market_cap,
                        price, shares_outstanding, risk_free_rate, eq_premium, growth_rate,
                        cost_of_capital, wealth_pc, fcff_value, terminal_value, share_value,
                        margin_of_safety, margin_of_safety_pc)
                    SELECT ticker, valuation_date, ent_name, industry, beta, market_cap,
                           price, shares_outstanding, risk_free_rate, eq_premium, growth_rate,
                           cost_of_capital, wealth_pc, fcff_value, terminal_value, share_value,
                           margin_of_safety, margin_of_safety_pc
                    FROM (
                        SELECT *, ROW_NUMBER() OVER (PARTITION BY ticker ORDER BY valuation_date DESC) rn
                        FROM valuation_old
                    ) WHERE rn = 1
                """)
                conn.execute("DROP TABLE valuation_old")
                conn.execute("COMMIT")
                logger.info("Migration complete.")
            except Exception as exc:
                try:
                    conn.execute("ROLLBACK")
                except Exception:
                    pass
                raise RuntimeError(f"Migration failed and was rolled back: {exc}") from exc
            finally:
                conn.isolation_level = orig_isolation
        else:
            conn.execute(schema_sql)
            conn.commit()
            logger.info("Table created successfully")
        # Add columns to existing tables that pre-date these fields
        for col_def in (
            "target_price REAL NOT NULL DEFAULT 0",
            "cik TEXT NOT NULL DEFAULT ''",
            "earnings_yield REAL NOT NULL DEFAULT 0",
            "dividend_yield REAL NOT NULL DEFAULT 0",
            "notes TEXT NOT NULL DEFAULT ''",
            "analyst_count INTEGER NOT NULL DEFAULT 0",
        ):
            try:
                conn.execute(f"ALTER TABLE valuation ADD COLUMN {col_def}")
                conn.commit()
                logger.info(f"Added column {col_def.split()[0]} to existing table")
            except sqlite3.OperationalError:
                pass  # column already exists
    except sqlite3.OperationalError as e:
        logger.warning(f"Failed to create tables: {e}")


def insert_valuation(conn, val):
    c = conn.cursor()
    c.execute(
        """INSERT OR REPLACE INTO valuation
        (ticker, valuation_date, ent_name, industry, cik, beta, market_cap, price,
            shares_outstanding, risk_free_rate, eq_premium, growth_rate,
            cost_of_capital, wealth_pc, fcff_value, terminal_value, share_value,
            margin_of_safety, margin_of_safety_pc, target_price, earnings_yield,
            dividend_yield, notes, analyst_count)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            val.ticker,
            val.valuation_date,
            val.ent_name,
            val.industry,
            val.cik,
            val.beta,
            val.market_cap,
            val.price,
            val.shares_outstanding,
            val.risk_free_rate,
            val.eq_premium,
            val.growth_rate,
            val.cost_of_capital,
            val.wealth_pc,
            val.fcff_value,
            val.terminal_value,
            val.share_value,
            val.margin_of_safety,
            val.margin_of_safety_pc,
            val.target_price,
            val.earnings_yield,
            val.dividend_yield,
            val.notes,
            val.analyst_count,
        ),
    )
    conn.commit()


def _rescore_tickers(db_path: str, tickers: list) -> None:
    """Re-score composite_scores for the given tickers immediately after valuation.

    Keeps composite scores in sync with valuations on a per-ticker basis so
    a full composite_score.py run is only needed after a bulk refresh.
    Errors are logged and swallowed — a scoring failure never aborts a valuation.
    """
    if not tickers:
        return
    try:
        from composite_score import (
            score_row, upsert, ensure_table as ensure_score_table,
        )
    except ImportError:
        logger.warning("composite_score not importable; skipping auto-rescore")
        return
    try:
        conn = sqlite3.connect(db_path, timeout=30)
        conn.row_factory = sqlite3.Row
        ensure_score_table(conn)
        moat_map = {r["ticker"]: dict(r)
                    for r in conn.execute("SELECT * FROM moat_scores").fetchall()}
        fd_map   = {r["ticker"]: dict(r)
                    for r in conn.execute("SELECT * FROM financial_data").fetchall()}
        for ticker in tickers:
            vrow = conn.execute(
                "SELECT * FROM valuation WHERE ticker=?", (ticker,)
            ).fetchone()
            if not vrow:
                continue
            r = score_row(dict(vrow), moat_map.get(ticker, {}), fd_map.get(ticker, {}))
            upsert(conn, r)
            logger.info(
                f"Composite score updated: {ticker} → {r['total_score']} ({r['designation']})"
            )
        conn.close()
        label = tickers[0] if len(tickers) == 1 else f"{len(tickers)} tickers"
        print(f"  Composite score(s) updated: {label}")
    except Exception as e:
        logger.warning(f"Composite score auto-update failed: {e}")


# ---------------------------------------------------------------------------
# Bank / financial-firm valuation  (FCFE equity DCF)
# ---------------------------------------------------------------------------


def _bank_payout_ratio(
    net_income: float, bv_equity_curr: float, bv_equity_prior: float, cash_flw: dict
) -> float:
    """
    Determine payout ratio for a bank using the most reliable source available.

    Priority:
      1. Dividends paid from cash flow (most direct; sum of quarterly outflows)
      2. Equity-change method (net income minus equity retained on balance sheet)
      3. Fallback: 40% payout (typical for well-run regional bank)

    AOCI swings (unrealized bond gains/losses) inflate the equity-change figure,
    so if that method would imply retention > 80% we prefer the dividend method.
    """
    payout = None

    # --- Method 1: actual dividends paid ---
    divs = sum(abs(v) for v in cash_flw.get("dividends_paid", []) if v)
    if net_income > 0 and divs > 0:
        payout_from_divs = divs / net_income
        if 0.05 <= payout_from_divs <= 0.95:
            payout = payout_from_divs
            logger.info(f"Payout ratio from dividends paid: {payout:.4f}")

    # --- Method 2: equity-change (only use if dividends unavailable/unreliable) ---
    if payout is None and net_income > 0:
        equity_change = bv_equity_curr - bv_equity_prior
        retention = equity_change / net_income
        if 0.0 <= retention <= 0.80:  # AOCI likely not distorting
            payout = 1.0 - retention
            logger.info(f"Payout ratio from equity change: {payout:.4f}")

    # --- Method 3: fallback ---
    if payout is None:
        payout = 0.40
        logger.info(f"Payout ratio: using fallback {payout:.4f}")

    return payout


def value_bank_stock(ticker: str, growth_period: int):
    """
    FCFE-based equity DCF for banks and financial firms.

    Key differences from FCFF:
    - Starts with Net Income, not EBIT
    - Reinvestment = equity retained to support asset growth (g / ROE)
    - Discounts at Cost of Equity, not WACC
    - No debt/cash adjustment — we work at the equity level throughout
    - No R&D capitalisation (not applicable to financials)
    """
    logger.info(f"Valuing {ticker} as financial firm (FCFE)")
    try:
        industry = hg_dcflib.get_industry(ticker)
        unlevered_beta = hg_dcflib.get_beta(industry)

        inc_stmnt = income_statement(ticker, MY_API_KEY)
        bal_sht = balance_sheet(ticker, MY_API_KEY, is_financial_or_reit=True)
        cash_flw = cash_flow_statement(ticker, MY_API_KEY)
        ent_quote = enterprise_quote(ticker, MY_API_KEY)

        valuation_date = str(date.today())
        price = ent_quote[0]
        shares_outstanding = ent_quote[1]
        market_cap = ent_quote[2]
        ent_name = ent_quote[3]
        dividend_yield = ent_quote[4]
        analyst_count = int(ent_quote[5])
        cik = hg_dcflib.get_cik(ticker)

        reported_net_income = inc_stmnt["netIncome"][0]
        if len(bal_sht["total_stockholders_equity"]) < 2:
            raise ValueError("Insufficient balance sheet history (need 2 years) for bank FCFE model")
        bv_equity_curr = bal_sht["total_stockholders_equity"][0]
        bv_equity_prior = bal_sht["total_stockholders_equity"][1]

        # Insurance firms: normalize NI over available years to smooth
        # underwriting cycles and catastrophe years.
        if is_insurance_firm(industry):
            net_income, ni_years = _normalized_net_income(inc_stmnt["netIncome"])
            logger.info(
                f"Insurance: normalized NI over {ni_years} years = {net_income:,.0f}  (TTM = {reported_net_income:,.0f})"
            )
        else:
            net_income = reported_net_income
            ni_years = 1

        # --- ROE and growth ---
        roe = net_income / bv_equity_curr if bv_equity_curr != 0 else 0.0
        logger.info(f"ROE = {roe:.4f}")

        payout_ratio = _bank_payout_ratio(
            reported_net_income, bv_equity_curr, bv_equity_prior, cash_flw
        )
        retention_ratio = 1.0 - payout_ratio
        growth_rate = _bank_growth_rate(roe, retention_ratio)
        logger.info(
            f"Growth rate = {growth_rate:.4f}  (ROE={roe:.4f} × retention={retention_ratio:.4f})"
        )

        # --- Cost of equity (no WACC — debt is operational for banks) ---
        bv_debt = calc_bv_debt(bal_sht)
        levered_beta = calc_levered_beta(unlevered_beta, bv_debt, market_cap, MARGINAL_TAX_RATE)
        cost_of_equity = RISK_FREE + (levered_beta * EQ_PREM)
        logger.info(f"Cost of Equity = {cost_of_equity:.4f}")

        # --- Project FCFE ---
        # Grow net income; FCFE = net income retained as dividends (payout fraction)
        payout_ratio = 1.0 - retention_ratio
        ni_n = []
        fcfe_n = []
        for year in range(growth_period):
            ni = net_income * (1 + growth_rate) ** (year + 1)
            ni_n.append(ni)
            fcfe_n.append(ni * payout_ratio)
            logger.info(f"FCFE year {year + 1} = {fcfe_n[-1]:,.2f}")

        fcfe_pv = sum(
            fcfe_n[y] / (1 + cost_of_equity) ** (y + 1) for y in range(growth_period)
        )

        # --- Stable phase ---
        # In stable phase ROE converges to cost of equity (competitive equilibrium)
        stable_beta = calc_stable_beta(unlevered_beta)
        stable_levered_beta = calc_levered_beta(
            stable_beta, bv_debt, market_cap, MARGINAL_TAX_RATE,
            de_cap=hg_dcflib.get_industry_de(industry),
        )
        stable_cost_of_equity = RISK_FREE + (stable_levered_beta * EQ_PREM)
        stable_growth = STABLE_GROWTH
        # stable ROE = stable CoE; shared helper guards non-positive denominators
        # and clamps to [0,1] — see docs/known_errors.md 2026-08-01.
        stable_reinv = calc_stable_reinvestment_rate(stable_growth, stable_cost_of_equity)
        stable_fcfe = fcfe_n[-1] * (1 + stable_growth) * (1 - stable_reinv)
        # Gordon Growth requires cost of equity > growth rate — see
        # docs/known_errors.md 2026-08-03 (CLMB/OSW/RCKY division-by-zero fix).
        if stable_cost_of_equity <= stable_growth:
            raise ValueError(
                f"Stable-phase cost of equity ({stable_cost_of_equity:.4f}) is at or "
                f"below the stable growth rate ({stable_growth:.4f}) — terminal value "
                f"is undefined (requires cost of equity > growth)."
            )
        terminal_value = stable_fcfe / (stable_cost_of_equity - stable_growth)
        terminal_value_pv = terminal_value / (1 + cost_of_equity) ** growth_period

        equity_value = fcfe_pv + terminal_value_pv
        intrinsic_value = equity_value / shares_outstanding  # both in consistent units

        safety_margin = float(intrinsic_value - price)
        safety_margin_pc = (
            1 - (price / intrinsic_value) if intrinsic_value != 0 else 0.0
        )
        wealth_pc = roe - cost_of_equity
        target_price = intrinsic_value * (1 + cost_of_equity)

        logger.info(f"Intrinsic value = {intrinsic_value:.2f}  Price = {price:.2f}")

        return Stock_Value(
            ticker=ticker,
            valuation_date=valuation_date,
            ent_name=ent_name,
            industry=industry,
            cik=cik,
            beta=levered_beta,
            market_cap=market_cap,
            price=price,
            shares_outstanding=shares_outstanding,
            risk_free_rate=RISK_FREE,
            eq_premium=EQ_PREM,
            growth_rate=growth_rate,
            cost_of_capital=cost_of_equity,  # equity rate, not WACC
            wealth_pc=wealth_pc,
            fcff_value=fcfe_pv,  # PV of FCFE
            terminal_value=terminal_value_pv,
            share_value=intrinsic_value,
            margin_of_safety=safety_margin,
            margin_of_safety_pc=safety_margin_pc,
            target_price=target_price,
            earnings_yield=0.0,  # FCFE model — EBIT/EV not applicable for banks
            dividend_yield=dividend_yield,
            analyst_count=analyst_count,
        )

    except Exception as e:
        logger.warning(f"Skipping {ticker}: {e}")
        logger.debug(traceback.format_exc())
        return None


# ---------------------------------------------------------------------------
# Single-stock valuation
# ---------------------------------------------------------------------------


def value_stock(ticker: str, growth_period: int, db_path: str | None = None):
    """
    Route to the correct valuation model based on industry:
      - REITs → skipped (FCFF/FCFE not applicable; FFO/AFFO model pending Phase 2)
      - Financial firms (banks, insurance, etc.) → FCFE equity DCF
      - All others → FCFF firm DCF
    """
    try:
        industry = hg_dcflib.get_industry(ticker)
    except Exception as e:
        logger.warning(f"Skipping {ticker}: {e}")
        return None

    if is_reit(industry):
        return value_reit_stock(ticker, growth_period)

    if is_financial_firm(industry):
        return value_bank_stock(ticker, growth_period)

    return _value_stock_fcff(ticker, growth_period, industry, db_path)


def _value_stock_fcff(ticker: str, growth_period: int, industry: str, db_path: str | None = None):
    """
    FCFF DCF valuation for non-financial firms.
    Returns a Stock_Value dataclass or None if any step fails.
    """
    logger.info(f"Valuing {ticker} ...")
    try:
        rd_years = hg_dcflib.get_rAndD_years(industry) + 1
        unlevered_beta = hg_dcflib.get_beta(industry)

        inc_stmnt = income_statement(ticker, MY_API_KEY)
        bal_sht = balance_sheet(ticker, MY_API_KEY)
        cash_flw = cash_flow_statement(ticker, MY_API_KEY)
        ent_quote = enterprise_quote(ticker, MY_API_KEY)

        valuation_date = str(date.today())
        price = ent_quote[0]
        shares_outstanding = ent_quote[1]
        market_cap = ent_quote[2]
        ent_name = ent_quote[3]
        dividend_yield = ent_quote[4]
        analyst_count = int(ent_quote[5])
        cik = hg_dcflib.get_cik(ticker)

        stable_beta = calc_stable_beta(unlevered_beta)
        eff_tax_rate = calc_tax_rate(inc_stmnt)
        fcff_data = calc_fcff(inc_stmnt, bal_sht, cash_flw, eff_tax_rate)

        ebiat = fcff_data[0]
        capex = fcff_data[1]
        chng_nc_wc = fcff_data[2]
        depreciation = fcff_data[3]

        amort_schedule = capitalizerAndD(ticker, rd_years, MY_API_KEY)
        logger.info(f"Amortization Schedule {amort_schedule}")

        adjusted_ebiat = calc_adj_ebiat(ebiat, amort_schedule)
        # Pre-tax adjusted EBIT — used as the base for projections so that
        # growth is applied to EBIT rather than EBIAT or FCFF.
        adjusted_ebit = (
            adjusted_ebiat / (1 - eff_tax_rate) if eff_tax_rate < 1 else adjusted_ebiat
        )
        firm_reinvestment = calc_reinvestment(
            capex, depreciation, chng_nc_wc, amort_schedule
        )
        adjusted_bv_equity = calc_adj_bv_equity(bal_sht, amort_schedule)
        bv_debt = calc_bv_debt(bal_sht)

        # Sanity check: if |EBIT| dwarfs the market cap by more than 10×,
        # the quarterly working-capital data is almost certainly corrupted.
        if market_cap > 0 and abs(adjusted_ebit) > 10 * market_cap:
            raise ValueError(
                f"Adjusted EBIT ({adjusted_ebit:,.0f}) is > 10× market cap "
                f"({market_cap:,.0f}) — likely bad WC data, skipping."
            )

        if adjusted_ebiat == 0:
            raise ValueError(
                f"Adjusted EBIAT is zero for {ticker} — cannot compute reinvestment rate."
            )
        reinvestment_rate = min(max(firm_reinvestment / adjusted_ebiat, 0.0), 1.0)
        logger.info(f"Reinvestment rate = {reinvestment_rate:,.4f}")

        return_on_capital, roc_notes = calc_gated_return_on_capital(
            ticker, adjusted_ebiat, adjusted_bv_equity, bv_debt, bal_sht, inc_stmnt, db_path
        )
        if return_on_capital is None:
            logger.warning(f"{ticker}: {roc_notes}")
            return Stock_Value(
                ticker=ticker, valuation_date=valuation_date, ent_name=ent_name,
                industry=industry, cik=cik,
                beta=calc_levered_beta(unlevered_beta, bv_debt, market_cap, MARGINAL_TAX_RATE),
                market_cap=market_cap,
                price=price, shares_outstanding=shares_outstanding,
                risk_free_rate=RISK_FREE, eq_premium=EQ_PREM,
                growth_rate=0.0, cost_of_capital=0.0, wealth_pc=0.0,
                fcff_value=0.0, terminal_value=0.0, share_value=0.0,
                margin_of_safety=0.0, margin_of_safety_pc=0.0, target_price=0.0,
                earnings_yield=0.0, dividend_yield=dividend_yield,
                notes=roc_notes,
                analyst_count=analyst_count,
            )
        growth_rate = min(calc_growth_rate(reinvestment_rate, return_on_capital), 0.30)

        levered_beta = calc_levered_beta(unlevered_beta, bv_debt, market_cap, MARGINAL_TAX_RATE)
        discount_rate = calc_discount_rate(
            inc_stmnt, bv_debt, market_cap, unlevered_beta, RISK_FREE, EQ_PREM
        )
        logger.info(f"disc rate {discount_rate:,.4f}")

        fcff_table = calc_expected_fcff(
            adjusted_ebit, eff_tax_rate, growth_rate, reinvestment_rate, growth_period
        )
        fcff_pv = calc_fcff_value(fcff_table, discount_rate, growth_period)

        terminal_cost_of_capital = calc_discount_rate(
            inc_stmnt, bv_debt, market_cap, stable_beta, RISK_FREE, EQ_PREM,
            de_cap=hg_dcflib.get_industry_de(industry),
        )
        ebit_last = adjusted_ebit * (1 + growth_rate) ** growth_period
        moat_weight = get_moat_weight(ticker, db_path)
        terminal_value_pv = calc_terminal_value(
            ebit_last,
            eff_tax_rate,
            terminal_cost_of_capital,
            discount_rate,
            STABLE_GROWTH,
            growth_period,
            moat_weight=moat_weight,
            explicit_roic=return_on_capital,
        )
        intrinsic_value = calc_intrinsic_value(
            fcff_pv,
            terminal_value_pv,
            bal_sht["cash_and_equivalents"][0],
            bv_debt,
            shares_outstanding,
        )

        safety_margin = float(intrinsic_value - price)
        logger.info(f"Safety Margin: {safety_margin:,.2f}")
        safety_margin_pc = (1 - (price / intrinsic_value)) if intrinsic_value != 0 else 0.0
        wealth_pc = return_on_capital - discount_rate
        target_price = intrinsic_value * (1 + discount_rate)
        _cash = bal_sht["cash_and_equivalents"][0]
        _ev = market_cap + bv_debt - _cash
        earnings_yield = adjusted_ebit / _ev if _ev > 0 else 0.0

        if return_on_capital > discount_rate:
            logger.info("Wealth Creator")
        else:
            logger.info("Wealth Destroyer")

        # Write the row even when the model produces a non-positive intrinsic
        # value rather than skipping it — a skip is indistinguishable in
        # valuation.db from "not yet attempted" or "rate-limited," which both
        # (a) erases why the ticker has no usable value and (b) causes every
        # future batch run to re-attempt it from scratch, burning an AV call
        # every night on a company whose negative EBIT isn't going to
        # un-happen tomorrow (same self-perpetuating-retry shape as the BK
        # bug in docs/known_errors.md). Flagging via `notes` instead lets a
        # query filter these out cheaply while preserving the audit trail —
        # see docs/known_errors.md 2026-07-31 "Negative FCFF intrinsic values".
        notes = (
            "Model produced non-positive intrinsic value — negative/deteriorating "
            "fundamentals; DCF result may not be economically meaningful"
            if intrinsic_value <= 0 else roc_notes
        )

        return Stock_Value(
            ticker=ticker,
            valuation_date=valuation_date,
            ent_name=ent_name,
            industry=industry,
            cik=cik,
            beta=levered_beta,
            market_cap=market_cap,
            price=price,
            shares_outstanding=shares_outstanding,
            risk_free_rate=RISK_FREE,
            eq_premium=EQ_PREM,
            growth_rate=growth_rate,
            cost_of_capital=discount_rate,
            wealth_pc=wealth_pc,
            fcff_value=fcff_pv,
            terminal_value=terminal_value_pv,
            share_value=intrinsic_value,
            margin_of_safety=safety_margin,
            margin_of_safety_pc=safety_margin_pc,
            target_price=target_price,
            earnings_yield=earnings_yield,
            dividend_yield=dividend_yield,
            analyst_count=analyst_count,
            notes=notes,
        )

    except Exception as e:
        logger.warning(f"Skipping {ticker}: {e}")
        logger.debug(traceback.format_exc())
        return None


# ---------------------------------------------------------------------------
# REIT valuation  (AFFO-based DDM)
# ---------------------------------------------------------------------------

def value_reit_stock(ticker: str, growth_period: int):
    """
    AFFO-based dividend discount model for REITs.

    Why not FCFF/FCFE:
    - REITs distribute 90%+ of income → retention ≈ 0 → FCFF growth = 0
    - Near-zero corporate tax distorts FCFF
    - Growth driven by acquisitions, not retained earnings

    Model:
    - AFFO = Net Income + D&A − CapEx  (AV CapEx = recurring, not acquisitions)
    - Payout = dividends_paid / AFFO
    - Growth = ROE × retention, capped at 15%
    - Discount at Cost of Equity (no WACC — REIT leverage is structural)
    """
    logger.info(f"Valuing {ticker} as REIT (AFFO DDM)")
    try:
        industry = hg_dcflib.get_industry(ticker)

        subtype_floor = reit_subtype_growth(ticker, industry)
        if subtype_floor is None:
            logger.warning(
                f"Skipping {ticker}: Mortgage REIT — AFFO DDM not applicable; manual review required"
            )
            return None

        unlevered_beta = hg_dcflib.get_beta(industry)

        inc_stmnt = income_statement(ticker, MY_API_KEY)
        bal_sht   = balance_sheet(ticker, MY_API_KEY, is_financial_or_reit=True)
        cash_flw  = cash_flow_statement(ticker, MY_API_KEY)
        ent_quote = enterprise_quote(ticker, MY_API_KEY)

        valuation_date    = str(date.today())
        price             = ent_quote[0]
        shares_outstanding = ent_quote[1]
        market_cap        = ent_quote[2]
        ent_name          = ent_quote[3]
        dividend_yield    = ent_quote[4]
        analyst_count     = int(ent_quote[5])
        cik               = hg_dcflib.get_cik(ticker)

        net_income     = inc_stmnt["netIncome"][0]
        da             = cash_flw["depreciation"][0] if cash_flw["depreciation"] else 0.0
        capex          = abs(cash_flw["capex"][0])           if cash_flw["capex"]          else 0.0
        dividends_paid = abs(cash_flw["dividends_paid"][0])  if cash_flw["dividends_paid"] else 0.0
        bv_equity      = bal_sht["total_stockholders_equity"][0]

        ffo  = net_income + da
        affo = max(ffo - capex, ffo * 0.80)  # floor at 80% of FFO to absorb edge cases

        payout_ratio   = min(dividends_paid / affo, 1.0) if affo > 0 else 0.85
        retention_ratio = 1.0 - payout_ratio

        roe             = net_income / bv_equity if bv_equity > 0 else 0.0
        # Use the higher of the retention-based rate and the sub-type floor.
        # The floor captures contractual lease escalators and structural growth
        # that exists independent of retained earnings (e.g. CPI escalators on
        # tower leases, 5G colocation, biological timber growth).
        growth_rate, retained_growth = _reit_growth_rate(roe, retention_ratio, subtype_floor)
        logger.info(
            f"AFFO={affo:,.0f}  payout={payout_ratio:.3f}  ROE={roe:.4f}  "
            f"retained_g={retained_growth:.4f}  subtype_floor={subtype_floor:.4f}  "
            f"g={growth_rate:.4f}"
        )

        bv_debt = calc_bv_debt(bal_sht)
        levered_beta = calc_levered_beta(unlevered_beta, bv_debt, market_cap, MARGINAL_TAX_RATE)
        cost_of_equity = RISK_FREE + (levered_beta * EQ_PREM)

        affo_n, div_n = [], []
        for year in range(growth_period):
            a = affo * (1 + growth_rate) ** (year + 1)
            affo_n.append(a)
            div_n.append(a * payout_ratio)

        div_pv = sum(div_n[y] / (1 + cost_of_equity) ** (y + 1) for y in range(growth_period))

        stable_beta             = calc_stable_beta(unlevered_beta)
        stable_levered_beta     = calc_levered_beta(stable_beta, bv_debt, market_cap, MARGINAL_TAX_RATE, de_cap=hg_dcflib.get_industry_de(industry))
        stable_cost_of_equity   = RISK_FREE + (stable_levered_beta * EQ_PREM)
        stable_growth           = STABLE_GROWTH
        # shared helper guards non-positive denominators and clamps to [0,1] —
        # see docs/known_errors.md 2026-08-01.
        stable_reinv            = calc_stable_reinvestment_rate(stable_growth, stable_cost_of_equity)
        stable_div              = div_n[-1] * (1 + stable_growth) * (1 - stable_reinv)
        # Gordon Growth requires cost of equity > growth rate — see
        # docs/known_errors.md 2026-08-03 (CLMB/OSW/RCKY division-by-zero fix).
        if stable_cost_of_equity <= stable_growth:
            raise ValueError(
                f"Stable-phase cost of equity ({stable_cost_of_equity:.4f}) is at or "
                f"below the stable growth rate ({stable_growth:.4f}) — terminal value "
                f"is undefined (requires cost of equity > growth)."
            )
        terminal_value          = stable_div / (stable_cost_of_equity - stable_growth)
        terminal_value_pv       = terminal_value / (1 + cost_of_equity) ** growth_period

        equity_value    = div_pv + terminal_value_pv
        intrinsic_value = equity_value / shares_outstanding
        # Write the row even when AFFO/growth-vs-CoE math produces a
        # non-positive IV rather than skipping — see the matching comment in
        # _value_stock_fcff() and docs/known_errors.md 2026-07-31 "Negative
        # FCFF intrinsic values" for why a skip is worse than a flagged row.
        notes = ""
        if intrinsic_value <= 0:
            logger.warning(
                f"{ticker}: AFFO model produced non-positive IV "
                f"({intrinsic_value:.2f}) — negative AFFO or growth ≥ CoE"
            )
            notes = (
                "AFFO model produced non-positive intrinsic value — negative "
                "AFFO or growth ≥ cost of equity; DCF result may not be "
                "economically meaningful"
            )
        safety_margin   = float(intrinsic_value - price)
        safety_margin_pc = (1 - price / intrinsic_value) if intrinsic_value != 0 else 0.0
        wealth_pc       = roe - cost_of_equity
        target_price    = intrinsic_value * (1 + cost_of_equity)

        logger.info(f"Intrinsic value = {intrinsic_value:.2f}  Price = {price:.2f}")

        return Stock_Value(
            ticker=ticker,
            valuation_date=valuation_date,
            ent_name=ent_name,
            industry=industry,
            cik=cik,
            beta=levered_beta,
            market_cap=market_cap,
            price=price,
            shares_outstanding=shares_outstanding,
            risk_free_rate=RISK_FREE,
            eq_premium=EQ_PREM,
            growth_rate=growth_rate,
            cost_of_capital=cost_of_equity,
            wealth_pc=wealth_pc,
            fcff_value=div_pv,
            terminal_value=terminal_value_pv,
            share_value=intrinsic_value,
            margin_of_safety=safety_margin,
            margin_of_safety_pc=safety_margin_pc,
            target_price=target_price,
            earnings_yield=0.0,
            dividend_yield=dividend_yield,
            analyst_count=analyst_count,
            notes=notes,
        )

    except Exception as e:
        logger.warning(f"Skipping {ticker}: {e}")
        logger.debug(traceback.format_exc())
        return None


# ---------------------------------------------------------------------------
# Single-stock detailed valuation (for Excel output)
# ---------------------------------------------------------------------------


def _stock_value_from_detail(d: dict) -> Stock_Value:
    """Build a Stock_Value dataclass from a value_stock_detail dict for DB insertion."""
    if d["model"] == "FCFE":
        cost_of_capital = d["cost_of_equity"]
        wealth_pc = d["roe"] - d["cost_of_equity"]
        fcff_value = d["fcfe_pv"]
    elif d["model"] == "AFFO":
        cost_of_capital = d["cost_of_equity"]
        wealth_pc = d["roe"] - d["cost_of_equity"]
        fcff_value = d["div_pv"]
    else:
        cost_of_capital = d["discount_rate"]
        wealth_pc = d["return_on_capital"] - d["discount_rate"]
        fcff_value = d["fcff_pv"]

    return Stock_Value(
        ticker=d["ticker"],
        valuation_date=d["valuation_date"],
        ent_name=d["ent_name"],
        industry=d["industry"],
        cik=d.get("cik", ""),
        beta=d["beta"],
        market_cap=d["market_cap"],
        price=d["price"],
        shares_outstanding=d["shares_outstanding"],
        risk_free_rate=d["risk_free"],
        eq_premium=d["eq_prem"],
        growth_rate=d["growth_rate"],
        cost_of_capital=cost_of_capital,
        wealth_pc=wealth_pc,
        fcff_value=fcff_value,
        terminal_value=d["terminal_value_pv"],
        share_value=d["intrinsic_value"],
        margin_of_safety=d["margin_of_safety"],
        margin_of_safety_pc=d["margin_of_safety_pc"],
        target_price=d["target_price"],
        earnings_yield=d.get("earnings_yield", 0.0),
        analyst_count=d.get("analyst_count", 0),
    )


def value_stock_detail(ticker: str, growth_period: int, db_path: str | None = None) -> dict | None:
    """
    Route to the correct detail valuation for the Excel report:
      - REITs          → skipped (FFO/AFFO model pending Phase 2)
      - Financial firms → FCFE bank detail
      - All others      → FCFF detail
    """
    try:
        industry = hg_dcflib.get_industry(ticker)
    except Exception as e:
        logger.warning(f"Skipping {ticker}: {e}")
        return None

    if is_reit(industry):
        return _value_reit_stock_detail(ticker, growth_period, industry)

    if is_financial_firm(industry):
        return _value_bank_stock_detail(ticker, growth_period, industry)
    return _value_stock_detail_fcff(ticker, growth_period, industry, db_path)


def _value_bank_stock_detail(
    ticker: str, growth_period: int, industry: str
) -> dict | None:
    """FCFE detail dict for bank/financial firms (used for Excel output)."""
    try:
        unlevered_beta = hg_dcflib.get_beta(industry)

        inc_stmnt = income_statement(ticker, MY_API_KEY)
        bal_sht = balance_sheet(ticker, MY_API_KEY, is_financial_or_reit=True)
        cash_flw = cash_flow_statement(ticker, MY_API_KEY)
        ent_quote = enterprise_quote(ticker, MY_API_KEY)

        price = ent_quote[0]
        shares_outstanding = ent_quote[1]
        market_cap = ent_quote[2]
        ent_name = ent_quote[3]
        analyst_count = int(ent_quote[5])
        cik = hg_dcflib.get_cik(ticker)

        reported_net_income = inc_stmnt["netIncome"][0]
        bv_equity_curr = bal_sht["total_stockholders_equity"][0]
        bv_equity_prior = bal_sht["total_stockholders_equity"][1]
        equity_change = bv_equity_curr - bv_equity_prior

        if is_insurance_firm(industry):
            net_income, ni_years = _normalized_net_income(inc_stmnt["netIncome"])
            logger.info(
                f"Insurance: normalized NI over {ni_years} years = {net_income:,.0f}  (TTM = {reported_net_income:,.0f})"
            )
        else:
            net_income = reported_net_income
            ni_years = 1

        roe = net_income / bv_equity_curr if bv_equity_curr != 0 else 0.0
        payout_ratio = _bank_payout_ratio(
            reported_net_income, bv_equity_curr, bv_equity_prior, cash_flw
        )
        retention_ratio = 1.0 - payout_ratio
        growth_rate = _bank_growth_rate(roe, retention_ratio)

        bv_debt = calc_bv_debt(bal_sht)
        levered_beta = calc_levered_beta(unlevered_beta, bv_debt, market_cap, MARGINAL_TAX_RATE)
        cost_of_equity = RISK_FREE + (levered_beta * EQ_PREM)

        ni_n, fcfe_n = [], []
        for year in range(growth_period):
            ni = net_income * (1 + growth_rate) ** (year + 1)
            ni_n.append(ni)
            fcfe_n.append(ni * payout_ratio)

        fcfe_pv = sum(
            fcfe_n[y] / (1 + cost_of_equity) ** (y + 1) for y in range(growth_period)
        )

        stable_beta = calc_stable_beta(unlevered_beta)
        stable_levered_beta = calc_levered_beta(stable_beta, bv_debt, market_cap, MARGINAL_TAX_RATE, de_cap=hg_dcflib.get_industry_de(industry))
        stable_cost_of_equity = RISK_FREE + (stable_levered_beta * EQ_PREM)
        stable_growth = STABLE_GROWTH
        # shared helper guards non-positive denominators and clamps to [0,1] —
        # see docs/known_errors.md 2026-08-01.
        stable_reinv = calc_stable_reinvestment_rate(stable_growth, stable_cost_of_equity)
        stable_fcfe = fcfe_n[-1] * (1 + stable_growth) * (1 - stable_reinv)
        # Gordon Growth requires cost of equity > growth rate — see
        # docs/known_errors.md 2026-08-03 (CLMB/OSW/RCKY division-by-zero fix).
        if stable_cost_of_equity <= stable_growth:
            raise ValueError(
                f"Stable-phase cost of equity ({stable_cost_of_equity:.4f}) is at or "
                f"below the stable growth rate ({stable_growth:.4f}) — terminal value "
                f"is undefined (requires cost of equity > growth)."
            )
        terminal_value_undiscounted = stable_fcfe / (
            stable_cost_of_equity - stable_growth
        )
        terminal_value_pv = (
            terminal_value_undiscounted / (1 + cost_of_equity) ** growth_period
        )

        equity_value = fcfe_pv + terminal_value_pv
        intrinsic_value = equity_value / shares_outstanding
        margin_of_safety = float(intrinsic_value - price)
        margin_of_safety_pc = (
            1 - (price / intrinsic_value) if intrinsic_value != 0 else 0
        )

        return {
            "model": "FCFE",
            "ticker": ticker,
            "ent_name": ent_name,
            "industry": industry,
            "cik": cik,
            "valuation_date": str(date.today()),
            # --- Inputs ---
            "net_income": net_income,
            "reported_net_income": reported_net_income,
            "ni_years": ni_years,
            "bv_equity": bv_equity_curr,
            "bv_equity_prior": bv_equity_prior,
            "equity_change": equity_change,
            "roe": roe,
            "retention_ratio": retention_ratio,
            "payout_ratio": payout_ratio,
            # --- Growth phase ---
            "growth_period": growth_period,
            "growth_rate": growth_rate,
            "beta": levered_beta,
            "stable_beta": stable_beta,
            "risk_free": RISK_FREE,
            "eq_prem": EQ_PREM,
            "cost_of_equity": cost_of_equity,
            # --- Year-by-year projections ---
            "ni_n": ni_n,
            "fcfe_n": fcfe_n,
            # --- Stable phase ---
            "stable_growth": stable_growth,
            "stable_reinv": stable_reinv,
            "stable_fcfe": stable_fcfe,
            "stable_cost_of_equity": stable_cost_of_equity,
            "terminal_value_undiscounted": terminal_value_undiscounted,
            "terminal_value_pv": terminal_value_pv,
            # --- Valuation ---
            "fcfe_pv": fcfe_pv,
            "equity_value": equity_value,
            "price": price,
            "shares_outstanding": shares_outstanding,
            "market_cap": market_cap,
            "intrinsic_value": intrinsic_value,
            "margin_of_safety": margin_of_safety,
            "margin_of_safety_pc": margin_of_safety_pc,
            "target_price": intrinsic_value * (1 + cost_of_equity),
        }

    except Exception as e:
        logger.warning(f"Skipping {ticker}: {e}")
        logger.debug(traceback.format_exc())
        return None


def _value_reit_stock_detail(
    ticker: str, growth_period: int, industry: str
) -> dict | None:
    """AFFO DDM detail dict for REITs (used for Excel output)."""
    try:
        unlevered_beta = hg_dcflib.get_beta(industry)

        inc_stmnt = income_statement(ticker, MY_API_KEY)
        bal_sht   = balance_sheet(ticker, MY_API_KEY, is_financial_or_reit=True)
        cash_flw  = cash_flow_statement(ticker, MY_API_KEY)
        ent_quote = enterprise_quote(ticker, MY_API_KEY)

        price              = ent_quote[0]
        shares_outstanding = ent_quote[1]
        market_cap         = ent_quote[2]
        ent_name           = ent_quote[3]
        cik                = hg_dcflib.get_cik(ticker)

        net_income     = inc_stmnt["netIncome"][0]
        da             = cash_flw["depreciation"][0] if cash_flw["depreciation"] else 0.0
        capex          = abs(cash_flw["capex"][0])           if cash_flw["capex"]          else 0.0
        dividends_paid = abs(cash_flw["dividends_paid"][0])  if cash_flw["dividends_paid"] else 0.0
        bv_equity      = bal_sht["total_stockholders_equity"][0]

        ffo  = net_income + da
        affo = max(ffo - capex, ffo * 0.80)

        payout_ratio    = min(dividends_paid / affo, 1.0) if affo > 0 else 0.85
        retention_ratio = 1.0 - payout_ratio
        roe             = net_income / bv_equity if bv_equity > 0 else 0.0
        subtype_floor   = reit_subtype_growth(ticker, industry)
        if subtype_floor is None:
            logger.warning(f"Skipping {ticker}: Mortgage REIT — AFFO DDM not applicable")
            return None
        growth_rate, retained_growth = _reit_growth_rate(roe, retention_ratio, subtype_floor)
        bv_debt         = calc_bv_debt(bal_sht)
        levered_beta    = calc_levered_beta(unlevered_beta, bv_debt, market_cap, MARGINAL_TAX_RATE)
        cost_of_equity  = RISK_FREE + (levered_beta * EQ_PREM)

        affo_n, div_n = [], []
        for year in range(growth_period):
            a = affo * (1 + growth_rate) ** (year + 1)
            affo_n.append(a)
            div_n.append(a * payout_ratio)

        div_pv = sum(div_n[y] / (1 + cost_of_equity) ** (y + 1) for y in range(growth_period))

        stable_beta           = calc_stable_beta(unlevered_beta)
        stable_levered_beta   = calc_levered_beta(stable_beta, bv_debt, market_cap, MARGINAL_TAX_RATE, de_cap=hg_dcflib.get_industry_de(industry))
        stable_cost_of_equity = RISK_FREE + (stable_levered_beta * EQ_PREM)
        stable_growth         = STABLE_GROWTH
        # shared helper guards non-positive denominators and clamps to [0,1] —
        # see docs/known_errors.md 2026-08-01.
        stable_reinv          = calc_stable_reinvestment_rate(stable_growth, stable_cost_of_equity)
        stable_div            = div_n[-1] * (1 + stable_growth) * (1 - stable_reinv)
        # Gordon Growth requires cost of equity > growth rate — see
        # docs/known_errors.md 2026-08-03 (CLMB/OSW/RCKY division-by-zero fix).
        if stable_cost_of_equity <= stable_growth:
            raise ValueError(
                f"Stable-phase cost of equity ({stable_cost_of_equity:.4f}) is at or "
                f"below the stable growth rate ({stable_growth:.4f}) — terminal value "
                f"is undefined (requires cost of equity > growth)."
            )
        terminal_value_undiscounted = stable_div / (stable_cost_of_equity - stable_growth)
        terminal_value_pv     = terminal_value_undiscounted / (1 + cost_of_equity) ** growth_period

        equity_value     = div_pv + terminal_value_pv
        intrinsic_value  = equity_value / shares_outstanding
        margin_of_safety = float(intrinsic_value - price)
        margin_of_safety_pc = (1 - price / intrinsic_value) if intrinsic_value != 0 else 0.0

        return {
            "model": "AFFO",
            "ticker": ticker,
            "ent_name": ent_name,
            "industry": industry,
            "cik": cik,
            "valuation_date": str(date.today()),
            # --- Inputs ---
            "net_income": net_income,
            "da": da,
            "capex": capex,
            "dividends_paid": dividends_paid,
            "ffo": ffo,
            "affo": affo,
            "bv_equity": bv_equity,
            "roe": roe,
            "payout_ratio": payout_ratio,
            "retention_ratio": retention_ratio,
            "retained_growth": retained_growth,
            "subtype_floor": subtype_floor,
            # --- Growth phase ---
            "growth_period": growth_period,
            "growth_rate": growth_rate,
            "beta": levered_beta,
            "stable_beta": stable_beta,
            "risk_free": RISK_FREE,
            "eq_prem": EQ_PREM,
            "cost_of_equity": cost_of_equity,
            # --- Year-by-year projections ---
            "affo_n": affo_n,
            "div_n": div_n,
            # --- Stable phase ---
            "stable_growth": stable_growth,
            "stable_reinv": stable_reinv,
            "stable_div": stable_div,
            "stable_cost_of_equity": stable_cost_of_equity,
            "terminal_value_undiscounted": terminal_value_undiscounted,
            "terminal_value_pv": terminal_value_pv,
            # --- Valuation ---
            "div_pv": div_pv,
            "equity_value": equity_value,
            "price": price,
            "shares_outstanding": shares_outstanding,
            "market_cap": market_cap,
            "intrinsic_value": intrinsic_value,
            "margin_of_safety": margin_of_safety,
            "margin_of_safety_pc": margin_of_safety_pc,
            "target_price": intrinsic_value * (1 + cost_of_equity),
        }

    except Exception as e:
        logger.warning(f"Skipping {ticker}: {e}")
        logger.debug(traceback.format_exc())
        return None


def _value_stock_detail_fcff(
    ticker: str, growth_period: int, industry: str, db_path: str | None = None
) -> dict | None:
    """FCFF detail dict for non-financial firms (used for Excel output)."""
    try:
        rd_years = hg_dcflib.get_rAndD_years(industry) + 1
        unlevered_beta = hg_dcflib.get_beta(industry)

        inc_stmnt = income_statement(ticker, MY_API_KEY)
        bal_sht = balance_sheet(ticker, MY_API_KEY)
        cash_flw = cash_flow_statement(ticker, MY_API_KEY)
        ent_quote = enterprise_quote(ticker, MY_API_KEY)

        price = ent_quote[0]
        shares_outstanding = ent_quote[1]
        market_cap = ent_quote[2]
        ent_name = ent_quote[3]
        cik = hg_dcflib.get_cik(ticker)

        stable_beta = calc_stable_beta(unlevered_beta)
        eff_tax_rate = calc_tax_rate(inc_stmnt)
        fcff_data = calc_fcff(inc_stmnt, bal_sht, cash_flw, eff_tax_rate)

        ebiat = fcff_data[0]
        capex = fcff_data[1]
        chng_nc_wc = fcff_data[2]
        depreciation = fcff_data[3]

        amort_schedule = capitalizerAndD(ticker, rd_years, MY_API_KEY)
        adjusted_ebiat = calc_adj_ebiat(ebiat, amort_schedule)
        adjusted_ebit = (
            adjusted_ebiat / (1 - eff_tax_rate) if eff_tax_rate < 1 else adjusted_ebiat
        )
        firm_reinvestment = calc_reinvestment(
            capex, depreciation, chng_nc_wc, amort_schedule
        )
        adjusted_bv_equity = calc_adj_bv_equity(bal_sht, amort_schedule)
        bv_debt = calc_bv_debt(bal_sht)

        if market_cap > 0 and abs(adjusted_ebit) > 10 * market_cap:
            raise ValueError(
                f"Adjusted EBIT ({adjusted_ebit:,.0f}) is > 10× market cap "
                f"({market_cap:,.0f}) — likely bad WC data, skipping."
            )

        if adjusted_ebiat == 0:
            raise ValueError(
                f"Adjusted EBIAT is zero for {ticker} — cannot compute reinvestment rate."
            )
        reinvestment_rate = min(max(firm_reinvestment / adjusted_ebiat, 0.0), 1.0)
        return_on_capital, roc_notes = calc_gated_return_on_capital(
            ticker, adjusted_ebiat, adjusted_bv_equity, bv_debt, bal_sht, inc_stmnt, db_path
        )
        if return_on_capital is None:
            raise ValueError(roc_notes)
        if roc_notes:
            logger.warning(f"{ticker}: {roc_notes}")
        growth_rate = min(calc_growth_rate(reinvestment_rate, return_on_capital), 0.30)

        # Compute discount rate components inline to capture intermediates
        levered_beta = calc_levered_beta(unlevered_beta, bv_debt, market_cap, MARGINAL_TAX_RATE)
        cost_of_equity = RISK_FREE + (levered_beta * EQ_PREM)
        try:
            int_cover = inc_stmnt["ebit"][0] / inc_stmnt["interest_expense"][0]
        except ZeroDivisionError:
            int_cover = 25
        def_spread = hg_dcflib.get_default_spread(int_cover)
        cost_of_debt_pretax = RISK_FREE + def_spread
        cost_of_debt_aftertax = cost_of_debt_pretax * (1 - MARGINAL_TAX_RATE)
        # WACC weights use market value of equity, book value of debt (Damodaran's
        # prescribed methodology — market debt is rarely observable, market equity
        # is trivial: price x shares). Previously weighted by book equity here,
        # diverging from calc_discount_rate() (used by the batch path, and by this
        # same function's own stable-phase rate two lines below) — see
        # docs/known_errors.md 2026-08-01.
        total_capital = market_cap + bv_debt
        percent_debt = bv_debt / total_capital if total_capital > 0 else 0.5
        percent_equity = 1 - percent_debt
        discount_rate = (cost_of_debt_aftertax * percent_debt) + (
            cost_of_equity * percent_equity
        )

        # Stable phase
        stable_levered_beta = calc_levered_beta(stable_beta, bv_debt, market_cap, MARGINAL_TAX_RATE, de_cap=hg_dcflib.get_industry_de(industry))
        stable_cost_of_equity = RISK_FREE + (stable_levered_beta * EQ_PREM)
        stable_cost_of_capital = calc_discount_rate(
            inc_stmnt, bv_debt, market_cap, stable_beta, RISK_FREE, EQ_PREM,
            de_cap=hg_dcflib.get_industry_de(industry),
        )
        stable_growth = STABLE_GROWTH
        moat_weight = get_moat_weight(ticker, db_path)
        if moat_weight:
            assumed_stable_roic = stable_cost_of_capital + moat_weight * (return_on_capital - stable_cost_of_capital)
            stable_reinv_rate = (
                min(max(stable_growth / assumed_stable_roic, 0.0), 1.0)
                if assumed_stable_roic > 0 else 1.0
            )
        else:
            stable_reinv_rate = calc_stable_reinvestment_rate(stable_growth, stable_cost_of_capital)

        # Year-by-year FCFF projections
        ebit_n, ebiat_n, reinv_n, fcff_n = [], [], [], []
        for year in range(growth_period):
            e = adjusted_ebit * (1 + growth_rate) ** (year + 1)
            eb = e * (1 - eff_tax_rate)
            r = eb * reinvestment_rate
            ebit_n.append(e)
            ebiat_n.append(eb)
            reinv_n.append(r)
            fcff_n.append(eb - r)

        fcff_pv = sum(
            fcff_n[y] / (1 + discount_rate) ** (y + 1) for y in range(growth_period)
        )

        # Terminal-year FCFF is recomputed at the stable-phase reinvestment
        # rate rather than carrying forward the explicit period's (typically
        # much higher) reinvestment rate — see calc_terminal_value() and
        # docs/known_errors.md 2026-07-31.
        # Display-only intermediates for the Excel report (stable_reinv_rate,
        # stable_fcff, terminal_value_undiscounted) — the authoritative
        # terminal_value_pv below comes from the shared calc_terminal_value(),
        # which recomputes the same values internally. See
        # docs/known_errors.md 2026-08-01 "FCFF terminal-value consolidation".
        terminal_ebit = ebit_n[-1] * (1 + stable_growth)
        terminal_ebiat = terminal_ebit * (1 - eff_tax_rate)
        stable_fcff = terminal_ebiat * (1 - stable_reinv_rate)
        # Gordon Growth requires cost of capital > growth rate. This block is
        # display-only (see comment above), but it still executes before the
        # authoritative calc_terminal_value() call below and would still
        # raise an unguarded ZeroDivisionError first — see docs/known_errors.md
        # 2026-08-03 (CLMB/OSW/RCKY division-by-zero fix).
        if stable_cost_of_capital <= stable_growth:
            raise ValueError(
                f"Stable-phase cost of capital ({stable_cost_of_capital:.4f}) is at or "
                f"below the stable growth rate ({stable_growth:.4f}) — terminal value "
                f"is undefined (requires cost of capital > growth)."
            )
        terminal_value_undiscounted = stable_fcff / (
            stable_cost_of_capital - stable_growth
        )
        terminal_value_pv = calc_terminal_value(
            ebit_n[-1], eff_tax_rate, stable_cost_of_capital, discount_rate,
            stable_growth, growth_period, moat_weight=moat_weight,
            explicit_roic=return_on_capital,
        )

        cash = bal_sht["cash_and_equivalents"][0]
        enterprise_value = fcff_pv + terminal_value_pv + cash - bv_debt
        intrinsic_value = enterprise_value / shares_outstanding
        margin_of_safety = float(intrinsic_value - price)
        margin_of_safety_pc = (
            1 - (price / intrinsic_value) if intrinsic_value != 0 else 0
        )

        # Non-cash working capital (current year)
        curr_nc_wc = (
            bal_sht["total_current_assets"][0] - bal_sht["cash_and_equivalents"][0]
        ) - (bal_sht["total_current_liabilities"][0] - bal_sht["short_term_debt"][0])
        revenue = inc_stmnt["totalRevenue"][0]
        wc_pct_revenue = curr_nc_wc / revenue if revenue != 0 else 0

        # Normalized valuation: if a single outlier quarter drove TTM EBIT negative,
        # rerun the model using TTM EBIT with that quarter excluded.  The result is
        # labeled "normalized" and shown alongside the GAAP figure in reports.
        ebit_anomaly = inc_stmnt.get("ebit_anomaly")
        norm_intrinsic_value = None
        norm_adjusted_ebit = None
        norm_return_on_capital = None
        norm_growth_rate = None
        if ebit_anomaly and adjusted_ebit < 0:
            norm_ttm_ebit_raw = ebit_anomaly["normalized_ttm_ebit"]
            norm_ebiat_raw = norm_ttm_ebit_raw * (1 - eff_tax_rate)
            norm_adjusted_ebiat = calc_adj_ebiat(norm_ebiat_raw, amort_schedule)
            if norm_adjusted_ebiat > 0:
                norm_adjusted_ebit = (
                    norm_adjusted_ebiat / (1 - eff_tax_rate) if eff_tax_rate < 1
                    else norm_adjusted_ebiat
                )
                norm_reinv_rate = min(
                    max(firm_reinvestment / norm_adjusted_ebiat, 0.0), 1.0
                )
                norm_return_on_capital, norm_roc_notes = calc_gated_return_on_capital(
                    ticker, norm_adjusted_ebiat, adjusted_bv_equity, bv_debt, bal_sht, inc_stmnt, db_path
                )
                if norm_return_on_capital is None:
                    # Same undefined-ROIC gate as the main path (see
                    # calc_gated_return_on_capital()) -- the normalized figure
                    # is a supplementary display value, not a hard-blocking
                    # path, so skip it gracefully rather than raising and
                    # losing the already-computed GAAP result above.
                    logger.warning(f"{ticker}: normalized valuation skipped -- {norm_roc_notes}")
                else:
                    if norm_roc_notes:
                        logger.warning(f"{ticker}: normalized valuation -- {norm_roc_notes}")
                    norm_growth_rate = min(
                        calc_growth_rate(norm_reinv_rate, norm_return_on_capital), 0.30
                    )
                    norm_fcff_table = calc_expected_fcff(
                        norm_adjusted_ebit, eff_tax_rate, norm_growth_rate,
                        norm_reinv_rate, growth_period
                    )
                    norm_fcff_pv = calc_fcff_value(norm_fcff_table, discount_rate, growth_period)
                    norm_ebit_last = norm_adjusted_ebit * (1 + norm_growth_rate) ** growth_period
                    norm_tv_pv = calc_terminal_value(
                        norm_ebit_last, eff_tax_rate, stable_cost_of_capital,
                        discount_rate, stable_growth, growth_period,
                        moat_weight=moat_weight, explicit_roic=norm_return_on_capital,
                    )
                    norm_ev = norm_fcff_pv + norm_tv_pv + cash - bv_debt
                    norm_intrinsic_value = norm_ev / shares_outstanding

        return {
            "model": "FCFF",
            "ticker": ticker,
            "ent_name": ent_name,
            "industry": industry,
            "cik": cik,
            "valuation_date": str(date.today()),
            # --- Inputs ---
            "normalized_ebit": inc_stmnt["ebit"][0],
            "adjusted_ebit": adjusted_ebit,
            "interest_expense": inc_stmnt["interest_expense"][0],
            "capex": capex,
            "depreciation": depreciation,
            "eff_tax_rate": eff_tax_rate,
            "revenue": revenue,
            "curr_nc_wc": curr_nc_wc,
            "chng_nc_wc": chng_nc_wc,
            "bv_debt": bv_debt,
            "adjusted_bv_equity": adjusted_bv_equity,
            "cash": cash,
            # --- Growth phase parameters ---
            "growth_period": growth_period,
            "growth_rate": growth_rate,
            "reinvestment_rate": reinvestment_rate,
            "return_on_capital": return_on_capital,
            "beta": levered_beta,
            "stable_beta": stable_beta,
            "risk_free": RISK_FREE,
            "eq_prem": EQ_PREM,
            "marginal_tax_rate": MARGINAL_TAX_RATE,
            # --- Discount rate components ---
            "cost_of_equity": cost_of_equity,
            "cost_of_debt_pretax": cost_of_debt_pretax,
            "cost_of_debt_aftertax": cost_of_debt_aftertax,
            "percent_debt": percent_debt,
            "percent_equity": percent_equity,
            "discount_rate": discount_rate,
            "wc_pct_revenue": wc_pct_revenue,
            # --- Stable phase ---
            "stable_growth": stable_growth,
            "stable_reinv_rate": stable_reinv_rate,
            "stable_fcff": stable_fcff,
            "stable_cost_of_equity": stable_cost_of_equity,
            "stable_cost_of_capital": stable_cost_of_capital,
            "terminal_value_undiscounted": terminal_value_undiscounted,
            "terminal_value_pv": terminal_value_pv,
            # --- Year-by-year projections ---
            "ebit_n": ebit_n,
            "ebiat_n": ebiat_n,
            "reinv_n": reinv_n,
            "fcff_n": fcff_n,
            # --- Valuation ---
            "fcff_pv": fcff_pv,
            "price": price,
            "shares_outstanding": shares_outstanding,
            "market_cap": market_cap,
            "intrinsic_value": intrinsic_value,
            "enterprise_value": enterprise_value,
            "margin_of_safety": margin_of_safety,
            "margin_of_safety_pc": margin_of_safety_pc,
            "target_price": intrinsic_value * (1 + discount_rate),
            "earnings_yield": adjusted_ebit / (market_cap + bv_debt - cash) if (market_cap + bv_debt - cash) > 0 else 0.0,
            # Anomaly detection — single outlier quarter driving negative TTM EBIT
            "ebit_anomaly": ebit_anomaly,
            "norm_intrinsic_value": norm_intrinsic_value,
            "norm_adjusted_ebit": norm_adjusted_ebit,
            "norm_return_on_capital": norm_return_on_capital,
            "norm_growth_rate": norm_growth_rate,
        }

    except Exception as e:
        logger.warning(f"Skipping {ticker}: {e}")
        logger.debug(traceback.format_exc())
        return None


# ---------------------------------------------------------------------------
# Excel single-stock report
# ---------------------------------------------------------------------------


def _generate_xlsx_bank(
    ws,
    d,
    gp,
    label,
    val_dollar,
    val_pct,
    section_header,
    BLUE_FILL,
    YELLOW_FILL,
    GREEN_FILL,
):
    """Populate the worksheet for a bank / financial firm (FCFE model)."""
    from openpyxl.styles import Font

    r = 1
    ws.cell(
        row=r, column=1, value=f"{d['ent_name']} ({d['ticker']}) — FCFE Valuation"
    ).font = Font(bold=True, size=13)
    r += 1
    ws.cell(
        row=r,
        column=1,
        value=f"Bank / Financial Firm  |  {d['valuation_date']}  |  Industry: {d['industry']}",
    )
    r += 1
    cik_val = d.get("cik", "")
    cik_cell = ws.cell(
        row=r,
        column=1,
        value=f"SEC EDGAR CIK: {cik_val}" if cik_val else "SEC EDGAR CIK: —",
    )
    cik_cell.font = Font(color="0563C1", underline="single")
    if cik_val:
        cik_cell.hyperlink = f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={cik_val}&type=10-K&dateb=&owner=include&count=10"
    r += 2

    # ---- Inputs --------------------------------------------------------
    section_header(r, 1, "Inputs")
    r += 1
    ni_label = (
        f"Net Income (normalized, {d['ni_years']}yr avg)"
        if d.get("ni_years", 1) > 1
        else "Net Income (TTM)"
    )
    inputs = [
        (ni_label, d["net_income"], "dollar"),
    ]
    if d.get("ni_years", 1) > 1:
        inputs.append(
            ("  Reported Net Income (TTM)", d["reported_net_income"], "dollar")
        )
    inputs += [
        ("Book Value of Equity (current)", d["bv_equity"], "dollar"),
        ("Book Value of Equity (prior yr)", d["bv_equity_prior"], "dollar"),
        ("Change in Book Equity", d["equity_change"], "dollar"),
        ("Return on Equity (ROE)", d["roe"], "pct"),
        ("Equity Retention Ratio", d["retention_ratio"], "pct"),
        ("Payout Ratio (FCFE / Net Income)", d["payout_ratio"], "pct"),
        ("Risk-free Rate", d["risk_free"], "pct"),
        ("Equity Risk Premium", d["eq_prem"], "pct"),
        ("Beta", d["beta"], "num"),
        ("Cost of Equity", d["cost_of_equity"], "pct"),
    ]
    for lbl, v, fmt in inputs:
        label(r, 1, lbl)
        if fmt == "dollar":
            val_dollar(r, 2, v, BLUE_FILL)
        elif fmt == "pct":
            val_pct(r, 2, v, BLUE_FILL)
        else:
            c = ws.cell(row=r, column=2, value=v)
            c.number_format = "0.00"
            c.fill = BLUE_FILL
        r += 1

    r += 1

    # ---- Parameters: High Growth vs Stable ----------------------------
    section_header(r, 1, "Parameters")
    label(r, 2, "High Growth", bold=True)
    label(r, 3, "Stable", bold=True)
    r += 1
    params = [
        ("Growth Rate", d["growth_rate"], "pct", d["stable_growth"]),
        ("Payout Ratio", d["payout_ratio"], "pct", 1 - d["stable_reinv"]),
        ("Beta", d["beta"], "num", d["stable_beta"]),
        ("Cost of Equity", d["cost_of_equity"], "pct", d["stable_cost_of_equity"]),
    ]
    for lbl, hg, fmt, st in params:
        label(r, 1, lbl)
        if fmt == "pct":
            val_pct(r, 2, hg, YELLOW_FILL)
            val_pct(r, 3, st, YELLOW_FILL)
        else:
            c = ws.cell(row=r, column=2, value=hg)
            c.number_format = "0.00"
            c.fill = YELLOW_FILL
            c2 = ws.cell(row=r, column=3, value=st)
            c2.number_format = "0.00"
        r += 1

    r += 1

    # ---- Year-by-year FCFE table --------------------------------------
    section_header(r, 1, f"Projected FCFE  (growth period = {gp} years)")
    for i in range(gp):
        ws.cell(row=r, column=2 + i, value=f"Year {i + 1}").font = Font(bold=True)
    r += 1

    table = [
        ("Expected Growth Rate", [d["growth_rate"]] * gp, "pct"),
        ("Net Income", d["ni_n"], "dollar"),
        ("FCFE (= NI × payout)", d["fcfe_n"], "dollar"),
        ("Cost of Equity", [d["cost_of_equity"]] * gp, "pct"),
        (
            "Cumulated CoE",
            [(1 + d["cost_of_equity"]) ** (y + 1) for y in range(gp)],
            "num",
        ),
        (
            "Present Value of FCFE",
            [d["fcfe_n"][y] / (1 + d["cost_of_equity"]) ** (y + 1) for y in range(gp)],
            "dollar",
        ),
    ]
    for lbl, values, fmt in table:
        label(r, 1, lbl)
        for i, v in enumerate(values):
            if fmt == "dollar":
                val_dollar(r, 2 + i, v)
            elif fmt == "pct":
                val_pct(r, 2 + i, v)
            else:
                c = ws.cell(row=r, column=2 + i, value=v)
                c.number_format = "0.0000"
        r += 1

    r += 1

    # ---- Stable phase -------------------------------------------------
    section_header(r, 1, "Stable Phase")
    r += 1
    stable = [
        ("Growth Rate in Stable Phase", d["stable_growth"], "pct"),
        ("Reinvestment Rate in Stable Phase", d["stable_reinv"], "pct"),
        ("FCFE in Stable Phase", d["stable_fcfe"], "dollar"),
        ("Cost of Equity in Stable Phase", d["stable_cost_of_equity"], "pct"),
        ("Terminal Value (undiscounted)", d["terminal_value_undiscounted"], "dollar"),
        ("PV of Terminal Value", d["terminal_value_pv"], "dollar"),
    ]
    for lbl, v, fmt in stable:
        label(r, 1, lbl)
        if fmt == "dollar":
            val_dollar(r, 2, v, YELLOW_FILL)
        else:
            val_pct(r, 2, v, YELLOW_FILL)
        r += 1

    r += 1

    # ---- Valuation summary --------------------------------------------
    section_header(r, 1, "Valuation")
    r += 1
    valuation = [
        ("PV of FCFE in High Growth Phase", d["fcfe_pv"], "dollar"),
        ("PV of Terminal Value", d["terminal_value_pv"], "dollar"),
        ("Equity Value", d["equity_value"], "dollar"),
        ("÷ Shares Outstanding", d["shares_outstanding"], "num"),
        ("Value of Equity per Share", d["intrinsic_value"], "dollar"),
        ("Target Price (1-yr)", d["target_price"], "dollar"),
        ("Stock Price", d["price"], "dollar"),
        ("Margin of Safety ($)", d["margin_of_safety"], "dollar"),
        ("Margin of Safety (%)", d["margin_of_safety_pc"], "pct"),
    ]
    for lbl, v, fmt in valuation:
        label(r, 1, lbl)
        fill = (
            GREEN_FILL
            if lbl
            in (
                "Value of Equity per Share",
                "Target Price (1-yr)",
                "Stock Price",
                "Margin of Safety ($)",
                "Margin of Safety (%)",
            )
            else None
        )
        if fmt == "dollar":
            val_dollar(r, 2, v, fill)
        elif fmt == "pct":
            val_pct(r, 2, v, fill)
        else:
            c = ws.cell(row=r, column=2, value=v)
            c.number_format = "#,##0"
            if fill:
                c.fill = fill
        r += 1


def _generate_xlsx_reit(
    ws, d, gp, label, val_dollar, val_pct, section_header,
    BLUE_FILL, YELLOW_FILL, GREEN_FILL,
):
    """Populate the worksheet for a REIT (AFFO DDM model)."""
    from openpyxl.styles import Font

    r = 1
    ws.cell(row=r, column=1,
            value=f"{d['ent_name']} ({d['ticker']}) — AFFO Valuation").font = Font(bold=True, size=13)
    r += 1
    ws.cell(row=r, column=1,
            value=f"REIT / AFFO DDM  |  {d['valuation_date']}  |  Industry: {d['industry']}")
    r += 1
    cik_val = d.get("cik", "")
    cik_cell = ws.cell(row=r, column=1,
                       value=f"SEC EDGAR CIK: {cik_val}" if cik_val else "SEC EDGAR CIK: —")
    cik_cell.font = Font(color="0563C1", underline="single")
    if cik_val:
        cik_cell.hyperlink = (
            f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany"
            f"&CIK={cik_val}&type=10-K&dateb=&owner=include&count=10"
        )
    r += 2

    # ---- Inputs --------------------------------------------------------
    section_header(r, 1, "Inputs")
    r += 1
    inputs = [
        ("Net Income (GAAP TTM)",            d["net_income"],      "dollar"),
        ("Depreciation & Amortization",      d["da"],              "dollar"),
        ("Recurring CapEx",                  d["capex"],           "dollar"),
        ("FFO  (Net Income + D&A)",          d["ffo"],             "dollar"),
        ("AFFO  (FFO − CapEx)",              d["affo"],            "dollar"),
        ("Dividends Paid",                   d["dividends_paid"],  "dollar"),
        ("Book Value of Equity",             d["bv_equity"],       "dollar"),
        ("Return on Equity (ROE)",           d["roe"],             "pct"),
        ("Payout Ratio (Dividends / AFFO)",  d["payout_ratio"],    "pct"),
        ("Equity Retention Ratio",           d["retention_ratio"], "pct"),
        ("Risk-free Rate",                   d["risk_free"],       "pct"),
        ("Equity Risk Premium",              d["eq_prem"],         "pct"),
        ("Beta",                             d["beta"],            "num"),
        ("Cost of Equity",                   d["cost_of_equity"],  "pct"),
    ]
    for lbl, v, fmt in inputs:
        label(r, 1, lbl)
        if fmt == "dollar":
            val_dollar(r, 2, v, BLUE_FILL)
        elif fmt == "pct":
            val_pct(r, 2, v, BLUE_FILL)
        else:
            c = ws.cell(row=r, column=2, value=v)
            c.number_format = "0.00"
            c.fill = BLUE_FILL
        r += 1

    r += 1

    # ---- Parameters: High Growth vs Stable ----------------------------
    section_header(r, 1, "Parameters")
    label(r, 2, "High Growth", bold=True)
    label(r, 3, "Stable", bold=True)
    r += 1
    params = [
        ("Growth Rate",    d["growth_rate"],  "pct", d["stable_growth"]),
        ("Payout Ratio",   d["payout_ratio"], "pct", 1 - d["stable_reinv"]),
        ("Beta",           d["beta"],         "num", d["stable_beta"]),
        ("Cost of Equity", d["cost_of_equity"], "pct", d["stable_cost_of_equity"]),
    ]
    for lbl, hg, fmt, st in params:
        label(r, 1, lbl)
        if fmt == "pct":
            val_pct(r, 2, hg, YELLOW_FILL)
            val_pct(r, 3, st, YELLOW_FILL)
        else:
            c = ws.cell(row=r, column=2, value=hg); c.number_format = "0.00"; c.fill = YELLOW_FILL
            c2 = ws.cell(row=r, column=3, value=st); c2.number_format = "0.00"; c2.fill = YELLOW_FILL
        r += 1

    r += 1

    # ---- Year-by-year AFFO projections --------------------------------
    section_header(r, 1, f"Projected Dividends  (growth period = {gp} years)")
    for col in range(1, gp + 1):
        label(r, col + 2, f"Year {col}", bold=True)
    r += 1

    rows_proj = [
        ("AFFO",            d["affo_n"]),
        ("Dividends",       d["div_n"]),
    ]
    for row_lbl, vals in rows_proj:
        label(r, 1, row_lbl)
        for col, v in enumerate(vals, 1):
            val_dollar(r, col + 2, v, None)
        r += 1

    # PV of dividends row
    label(r, 1, "PV of Dividends (sum)")
    val_dollar(r, 2, d["div_pv"], GREEN_FILL)
    r += 2

    # ---- Stable phase -------------------------------------------------
    section_header(r, 1, "Stable Phase")
    r += 1
    stable_rows = [
        ("Growth Rate in Stable Phase",          d["stable_growth"],                "pct"),
        ("Reinvestment Rate in Stable Phase",     d["stable_reinv"],                 "pct"),
        ("Cost of Equity in Stable Phase",        d["stable_cost_of_equity"],        "pct"),
        ("Stable Dividend",                       d["stable_div"],                   "dollar"),
        ("Terminal Value (undiscounted)",          d["terminal_value_undiscounted"],  "dollar"),
        ("PV of Terminal Value",                  d["terminal_value_pv"],            "dollar"),
    ]
    for lbl, v, fmt in stable_rows:
        label(r, 1, lbl)
        if fmt == "dollar":
            val_dollar(r, 2, v, BLUE_FILL)
        else:
            val_pct(r, 2, v, BLUE_FILL)
        r += 1

    r += 1

    # ---- Valuation summary --------------------------------------------
    section_header(r, 1, "Valuation Summary")
    r += 1
    val_rows = [
        ("PV of Dividends (growth period)",  d["div_pv"],            "dollar"),
        ("PV of Terminal Value",             d["terminal_value_pv"], "dollar"),
        ("Equity Value",                     d["equity_value"],      "dollar"),
        ("Intrinsic Value / Share",          d["intrinsic_value"],   "dollar"),
        ("Current Price",                    d["price"],             "dollar"),
        ("Margin of Safety ($)",             d["margin_of_safety"],  "dollar"),
        ("Margin of Safety (%)",             d["margin_of_safety_pc"], "pct"),
    ]
    for lbl, v, fmt in val_rows:
        label(r, 1, lbl)
        if fmt == "dollar":
            val_dollar(r, 2, v, GREEN_FILL)
        else:
            val_pct(r, 2, v, GREEN_FILL)
        r += 1


def generate_xlsx(d: dict, output_path: str) -> None:
    """Write a Damodaran-style FCFF valuation worksheet for a single stock."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = d["ticker"]

    # ---- helpers --------------------------------------------------------
    BLUE_FILL = PatternFill("solid", fgColor="DDEEFF")
    YELLOW_FILL = PatternFill("solid", fgColor="FFFACD")
    GREEN_FILL = PatternFill("solid", fgColor="D4EDDA")
    HEADER_FONT = Font(bold=True)

    FMT_DOLLAR = "#,##0.00"
    FMT_PCT = "0.00%"

    def label(row, col, text, bold=False):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=bold)
        return c

    def val_dollar(row, col, v, fill=None):
        c = ws.cell(row=row, column=col, value=v)
        c.number_format = FMT_DOLLAR
        c.alignment = Alignment(horizontal="right")
        if fill:
            c.fill = fill
        return c

    def val_pct(row, col, v, fill=None):
        c = ws.cell(row=row, column=col, value=v)
        c.number_format = FMT_PCT
        c.alignment = Alignment(horizontal="right")
        if fill:
            c.fill = fill
        return c

    def section_header(row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=True, underline="single")
        return c

    gp = d["growth_period"]

    # ---- Column widths --------------------------------------------------
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    for col in range(4, 4 + gp):
        ws.column_dimensions[get_column_letter(col)].width = 12

    if d.get("model") == "AFFO":
        _generate_xlsx_reit(
            ws, d, gp, label, val_dollar, val_pct, section_header,
            BLUE_FILL, YELLOW_FILL, GREEN_FILL,
        )
        wb.save(output_path)
        print(f"Saved: {output_path}")
        return

    if d.get("model") == "FCFE":
        _generate_xlsx_bank(
            ws,
            d,
            gp,
            label,
            val_dollar,
            val_pct,
            section_header,
            BLUE_FILL,
            YELLOW_FILL,
            GREEN_FILL,
        )
        wb.save(output_path)
        print(f"Saved: {output_path}")
        return

    _generate_xlsx_fcff(
        ws, d, gp, label, val_dollar, val_pct, section_header,
        BLUE_FILL, YELLOW_FILL, GREEN_FILL,
    )
    wb.save(output_path)
    print(f"Saved: {output_path}")


def _generate_xlsx_fcff(
    ws, d, gp, label, val_dollar, val_pct, section_header,
    BLUE_FILL, YELLOW_FILL, GREEN_FILL,
):
    """Populate the worksheet for a standard non-financial firm (FCFF model)."""
    from openpyxl.styles import Font

    HEADER_FONT = Font(bold=True)

    # ====================================================================
    # TITLE
    # ====================================================================
    r = 1
    ws.cell(row=r, column=1, value=f"{d['ent_name']} ({d['ticker']})").font = Font(
        bold=True, size=13
    )
    r += 1
    ws.cell(
        row=r,
        column=1,
        value=f"FCFF Valuation  |  {d['valuation_date']}  |  Industry: {d['industry']}",
    )
    r += 1
    cik_val = d.get("cik", "")
    cik_cell = ws.cell(
        row=r,
        column=1,
        value=f"SEC EDGAR CIK: {cik_val}" if cik_val else "SEC EDGAR CIK: —",
    )
    cik_cell.font = Font(color="0563C1", underline="single")
    if cik_val:
        cik_cell.hyperlink = f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={cik_val}&type=10-K&dateb=&owner=include&count=10"
    r += 2

    # ====================================================================
    # SECTION 1 — INPUTS
    # ====================================================================
    section_header(r, 1, "Inputs")
    r += 1

    inputs = [
        ("Normalized EBIT (before adjustments)", d["normalized_ebit"], "dollar"),
        ("Adjusted EBIT", d["adjusted_ebit"], "dollar"),
        ("Adjusted Interest Expense", d["interest_expense"], "dollar"),
        ("Adjusted Capital Spending (avg 5yr)", d["capex"], "dollar"),
        ("Adjusted Depreciation & Amort'n", d["depreciation"], "dollar"),
        ("Tax Rate on Income", d["eff_tax_rate"], "pct"),
        ("Current Revenues", d["revenue"], "dollar"),
        ("Current Non-cash Working Capital", d["curr_nc_wc"], "dollar"),
        ("Chg. Working Capital", d["chng_nc_wc"], "dollar"),
        ("Adjusted Book Value of Debt", d["bv_debt"], "dollar"),
        ("Adjusted Book Value of Equity", d["adjusted_bv_equity"], "dollar"),
    ]
    for lbl, v, fmt in inputs:
        label(r, 1, lbl)
        if fmt == "dollar":
            val_dollar(r, 2, v, BLUE_FILL)
        else:
            val_pct(r, 2, v, BLUE_FILL)
        r += 1

    r += 1  # blank

    # ====================================================================
    # SECTION 2 — PARAMETERS: HIGH GROWTH vs STABLE
    # ====================================================================
    section_header(r, 1, "Parameters")
    label(r, 2, "High Growth", bold=True)
    label(r, 3, "Stable", bold=True)
    r += 1

    params = [
        ("Length of High Growth Period", d["growth_period"], "int", "Forever"),
        ("Growth Rate", d["growth_rate"], "pct", d["stable_growth"]),
        ("Beta used for stock", d["beta"], "num", d["stable_beta"]),
        ("Risk-free Rate", d["risk_free"], "pct", d["risk_free"]),
        ("Equity Risk Premium", d["eq_prem"], "pct", d["eq_prem"]),
        (
            "Pre-tax Cost of Debt",
            d["cost_of_debt_pretax"],
            "pct",
            d["cost_of_debt_pretax"],
        ),
        (
            "Effective Tax Rate (cash flow)",
            d["eff_tax_rate"],
            "pct",
            d["marginal_tax_rate"],
        ),
        (
            "Marginal Tax Rate (cost of debt)",
            d["marginal_tax_rate"],
            "pct",
            d["marginal_tax_rate"],
        ),
        ("Return on Capital", d["return_on_capital"], "pct", d["return_on_capital"]),
        ("Reinvestment Rate", d["reinvestment_rate"], "pct", d["stable_reinv_rate"]),
        ("Debt / (Debt + Equity)", d["percent_debt"], "pct", d["percent_debt"]),
    ]
    for lbl, hg_val, fmt, st_val in params:
        label(r, 1, lbl)
        if fmt == "pct":
            val_pct(r, 2, hg_val, YELLOW_FILL)
            val_pct(r, 3, st_val, YELLOW_FILL) if isinstance(
                st_val, float
            ) else ws.cell(row=r, column=3, value=st_val)
        elif fmt == "int":
            ws.cell(row=r, column=2, value=hg_val).fill = YELLOW_FILL
            ws.cell(row=r, column=3, value=st_val)
        else:
            c = ws.cell(row=r, column=2, value=hg_val)
            c.number_format = "0.00"
            c.fill = YELLOW_FILL
            c2 = ws.cell(row=r, column=3, value=st_val)
            c2.number_format = "0.00"
        r += 1

    r += 1  # blank

    # ====================================================================
    # SECTION 3 — COST OF CAPITAL OUTPUT
    # ====================================================================
    section_header(r, 1, "Cost of Capital — Output")
    r += 1

    coc_rows = [
        ("Cost of Equity", d["cost_of_equity"], "pct"),
        ("Equity / (Debt + Equity)", d["percent_equity"], "pct"),
        ("After-tax Cost of Debt", d["cost_of_debt_aftertax"], "pct"),
        ("Debt / (Debt + Equity)", d["percent_debt"], "pct"),
        ("Cost of Capital (WACC)", d["discount_rate"], "pct"),
    ]
    for lbl, v, fmt in coc_rows:
        label(r, 1, lbl)
        val_pct(r, 2, v, GREEN_FILL)
        r += 1

    r += 1

    label(r, 1, "Working Capital as % of Revenue")
    val_pct(r, 2, d["wc_pct_revenue"])
    r += 2

    # ====================================================================
    # SECTION 4 — YEAR-BY-YEAR FCFF TABLE
    # ====================================================================
    section_header(r, 1, f"Projected FCFF  (growth period = {gp} years)")
    year_cols = list(range(1, gp + 1))
    for i, yr in enumerate(year_cols):
        ws.cell(row=r, column=2 + i, value=f"Year {yr}").font = HEADER_FONT
    r += 1

    net_capex = d["capex"] - d["depreciation"]
    # Proportion of reinvestment attributable to net capex vs WC change
    total_reinv0 = (
        net_capex + d["chng_nc_wc"] if (net_capex + d["chng_nc_wc"]) != 0 else 1
    )
    capex_frac = net_capex / total_reinv0
    wc_frac = d["chng_nc_wc"] / total_reinv0

    table_rows = [
        ("Expected Growth Rate", [d["growth_rate"]] * gp, "pct"),
        (
            "Cumulated Growth",
            [(1 + d["growth_rate"]) ** (y + 1) - 1 for y in range(gp)],
            "pct",
        ),
        ("Reinvestment Rate", [d["reinvestment_rate"]] * gp, "pct"),
        ("EBIT", d["ebit_n"], "dollar"),
        ("Tax Rate (cash flow)", [d["eff_tax_rate"]] * gp, "pct"),
        ("EBIT × (1 − tax rate)", d["ebiat_n"], "dollar"),
        (
            "− (CapEx − Depreciation)",
            [-r_v * capex_frac for r_v in d["reinv_n"]],
            "dollar",
        ),
        ("− Chg. Working Capital", [-r_v * wc_frac for r_v in d["reinv_n"]], "dollar"),
        ("Free Cash Flow to Firm", d["fcff_n"], "dollar"),
        ("Cost of Capital", [d["discount_rate"]] * gp, "pct"),
        (
            "Cumulated Cost of Capital",
            [(1 + d["discount_rate"]) ** (y + 1) for y in range(gp)],
            "num",
        ),
        (
            "Present Value",
            [d["fcff_n"][y] / (1 + d["discount_rate"]) ** (y + 1) for y in range(gp)],
            "dollar",
        ),
    ]

    for lbl, values, fmt in table_rows:
        label(r, 1, lbl)
        for i, v in enumerate(values):
            if fmt == "dollar":
                val_dollar(r, 2 + i, v)
            elif fmt == "pct":
                val_pct(r, 2 + i, v)
            else:
                c = ws.cell(row=r, column=2 + i, value=v)
                c.number_format = "0.0000"
        r += 1

    r += 1

    # ====================================================================
    # SECTION 5 — STABLE PHASE
    # ====================================================================
    section_header(r, 1, "Stable Phase")
    r += 1

    stable_rows = [
        ("Growth Rate in Stable Phase", d["stable_growth"], "pct"),
        ("Reinvestment Rate in Stable Phase", d["stable_reinv_rate"], "pct"),
        ("FCFF in Stable Phase", d["stable_fcff"], "dollar"),
        ("Cost of Equity in Stable Phase", d["stable_cost_of_equity"], "pct"),
        ("Cost of Capital in Stable Phase", d["stable_cost_of_capital"], "pct"),
        (
            "Value at End of Growth Phase (TV)",
            d["terminal_value_undiscounted"],
            "dollar",
        ),
        ("PV of Terminal Value", d["terminal_value_pv"], "dollar"),
    ]
    for lbl, v, fmt in stable_rows:
        label(r, 1, lbl)
        if fmt == "dollar":
            val_dollar(r, 2, v, YELLOW_FILL)
        else:
            val_pct(r, 2, v, YELLOW_FILL)
        r += 1

    r += 1

    # ====================================================================
    # SECTION 6 — VALUATION SUMMARY
    # ====================================================================
    section_header(r, 1, "Valuation")
    r += 1

    valuation_rows = [
        ("PV of FCFF in High Growth Phase", d["fcff_pv"], "dollar"),
        ("PV of Terminal Value of Firm", d["terminal_value_pv"], "dollar"),
        ("Value of Operating Assets", d["fcff_pv"] + d["terminal_value_pv"], "dollar"),
        ("+ Cash & Non-operating Assets", d["cash"], "dollar"),
        ("Value of Firm", d["enterprise_value"] + d["bv_debt"], "dollar"),
        ("− Market Value of Debt", d["bv_debt"], "dollar"),
        ("Market Value of Equity", d["enterprise_value"], "dollar"),
        ("÷ Shares Outstanding (000s)", d["shares_outstanding"], "num"),
        ("Value of Equity per Share", d["intrinsic_value"], "dollar"),
        ("Target Price (1-yr)", d["target_price"], "dollar"),
        ("Stock Price", d["price"], "dollar"),
        ("Margin of Safety ($)", d["margin_of_safety"], "dollar"),
        ("Margin of Safety (%)", d["margin_of_safety_pc"], "pct"),
    ]
    for lbl, v, fmt in valuation_rows:
        label(r, 1, lbl)
        fill = (
            GREEN_FILL
            if lbl
            in (
                "Value of Equity per Share",
                "Target Price (1-yr)",
                "Stock Price",
                "Margin of Safety ($)",
                "Margin of Safety (%)",
            )
            else None
        )
        if fmt == "dollar":
            val_dollar(r, 2, v, fill)
        elif fmt == "pct":
            val_pct(r, 2, v, fill)
        else:
            c = ws.cell(row=r, column=2, value=v)
            c.number_format = "#,##0"
            if fill:
                c.fill = fill
        r += 1


# ---------------------------------------------------------------------------
# XLSX report generation
# ---------------------------------------------------------------------------


def generate_summary_xlsx(
    valuations: list, output_path: str, index_label: str = "S&P 500"
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation Summary"

    DARK_FILL = PatternFill("solid", fgColor="343A40")
    GREEN_FILL = PatternFill("solid", fgColor="D4EDDA")
    YELLOW_FILL = PatternFill("solid", fgColor="FFF3CD")
    RED_FILL = PatternFill("solid", fgColor="F8D7DA")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Metadata rows
    today = date.today().strftime("%B %d, %Y")
    ws.cell(
        row=1, column=1, value=f"{index_label} FCFF Valuation — {today}"
    ).font = Font(bold=True, size=13)
    ws.cell(
        row=2,
        column=1,
        value=f"Risk-free rate: {RISK_FREE * 100:.2f}%  |  ERP: {EQ_PREM * 100:.2f}%  |  "
        f"Stable growth: {STABLE_GROWTH * 100:.1f}%  |  "
        f"Growth period: {GROWTH_PERIOD} yrs  |  Stocks valued: {len(valuations)}",
    ).font = Font(italic=True, color="666666")

    # Header row
    headers = [
        "Ticker",
        "CIK",
        "Company",
        "Industry",
        "Price",
        "Intrinsic Value",
        "Target Price\n(1-yr)",
        "MoS ($)",
        "MoS (%)",
        "Growth Rate",
        "Cost of Capital",
        "Excess Return\n(ROIC-WACC)",
        "Unlevered Beta",
        "Market Cap ($B)",
    ]
    HDR_ROW = 4
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=HDR_ROW, column=col, value=hdr)
        c.font = HEADER_FONT
        c.fill = DARK_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER

    # Column widths / number formats shared by both tabs
    col_widths = [8, 12, 28, 20, 10, 14, 14, 10, 10, 12, 14, 16, 13, 14]
    num_fmts = [
        None,
        None,
        None,
        None,
        '"$"#,##0.00',
        '"$"#,##0.00',
        '"$"#,##0.00',
        '"$"#,##0.00',
        "0.0%",
        "0.0%",
        "0.0%",
        "0.0%",
        "0.000",
        "#,##0.00",
    ]

    def _write_tab(sheet, rows):
        """Header row + data rows + column widths + freeze pane for one tab
        — shared by the main Valuation Summary and Value Creators sheets."""
        for col, hdr in enumerate(headers, 1):
            c = sheet.cell(row=HDR_ROW, column=col, value=hdr)
            c.font = HEADER_FONT
            c.fill = DARK_FILL
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = BORDER
        sheet.row_dimensions[HDR_ROW].height = 30

        for row_idx, v in enumerate(rows, HDR_ROW + 1):
            if v.margin_of_safety >= 1:
                row_fill = GREEN_FILL
            elif v.margin_of_safety >= 0:
                row_fill = YELLOW_FILL
            else:
                row_fill = RED_FILL

            row_data = [
                v.ticker,
                v.cik,
                v.ent_name,
                v.industry,
                v.price,
                v.share_value,
                v.target_price,
                v.margin_of_safety,
                v.margin_of_safety_pc,
                v.growth_rate,
                v.cost_of_capital,
                v.wealth_pc,
                v.beta,
                v.market_cap / 1e9,
            ]
            for col, (val, fmt) in enumerate(zip(row_data, num_fmts), 1):
                c = sheet.cell(row=row_idx, column=col, value=val)
                c.fill = row_fill
                c.border = BORDER
                if fmt:
                    c.number_format = fmt
                c.alignment = Alignment(horizontal="left" if col <= 4 else "right")

        for col, w in enumerate(col_widths, 1):
            sheet.column_dimensions[get_column_letter(col)].width = w
        sheet.freeze_panes = sheet.cell(row=HDR_ROW + 1, column=1)

    _write_tab(ws, valuations)

    # ----------------------------------------------------------------
    # Tab 2 — Value Creators with positive MoS, sorted by Excess Return desc
    # ----------------------------------------------------------------
    creators = sorted(
        [v for v in valuations if v.wealth_pc > 0 and v.margin_of_safety > 0],
        key=lambda v: v.wealth_pc,
        reverse=True,
    )

    vc = wb.create_sheet(title="Value Creators")

    vc.cell(
        row=1,
        column=1,
        value=f"{index_label} — Value Creators & Undervalued Stocks — {today}",
    ).font = Font(bold=True, size=13)
    vc.cell(
        row=2,
        column=1,
        value=f"{len(creators)} of {len(valuations)} stocks (ROIC > WACC or MoS > $0) — sorted by Excess Return (ROIC − WACC) descending",
    ).font = Font(italic=True, color="666666")

    _write_tab(vc, creators)

    wb.save(output_path)
    print(f"Saved: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def _write_market_data_status(erp_status: dict, risk_free_status: dict) -> None:
    """
    Record refresh_market_data()'s outcome to data/market_data_fetch_status.json
    so Iggy's nightly valuation report can surface a fetch failure to Jim (see
    iggy-valuation-update SKILL.md's "Notes for Addie" section).

    get_erp()/get_risk_free() have no timeout or retry (unlike hg_dcflib's AV
    fetchers) — a hang or failure here previously only produced a logger.warning
    line nobody would see until this file existed. Best-effort: a failure to
    write this status file is logged but never blocks a valuation run.
    """
    status_path = _Path(os.path.abspath(__file__)).parent.parent / "data" / "market_data_fetch_status.json"
    try:
        status_path.parent.mkdir(parents=True, exist_ok=True)
        with open(status_path, "w") as f:
            json.dump({
                "date": date.today().isoformat(),
                "erp": erp_status,
                "risk_free": risk_free_status,
                "eq_prem_used": EQ_PREM,
                "risk_free_used": RISK_FREE,
            }, f, indent=2)
    except Exception as e:
        logger.warning(f"Could not write market_data_fetch_status.json: {e}")


def refresh_market_data():
    """
    Fetch live ERP and risk-free rate, falling back to the module-level
    constants on any failure. Deferred from module level so a slow/failed
    network call doesn't block argument parsing or import. Callers that
    import this module as a library (e.g. stock_analysis.py) must call
    this explicitly — it does not run automatically except via main().

    Always records the outcome via _write_market_data_status(), even on
    success — see that function's docstring for why.
    """
    global EQ_PREM, RISK_FREE
    erp_status = {"ok": True, "error": None}
    risk_free_status = {"ok": True, "error": None}
    try:
        _erp = hg_dcflib.get_erp()
        if _erp is None:
            erp_status = {"ok": False, "error": "get_erp() returned None (no Implied ERP % parsed)"}
            logger.warning(f"ERP returned None; using fallback {EQ_PREM:.4f}")
        else:
            EQ_PREM = _erp
            logger.info(f"ERP: {EQ_PREM:.4f}")
    except Exception as e:
        erp_status = {"ok": False, "error": str(e)}
        logger.warning(f"ERP fetch failed ({e}); using fallback {EQ_PREM:.4f}")
    try:
        _rf = hg_dcflib.get_risk_free(FRED_KEY)
        if _rf is None:
            risk_free_status = {"ok": False, "error": "get_risk_free() returned None (non-200 FRED response)"}
            logger.warning(f"Risk-free rate returned None; using fallback {RISK_FREE:.4f}")
        else:
            RISK_FREE = _rf
            logger.info(f"Risk-free: {RISK_FREE:.4f}")
    except Exception as e:
        risk_free_status = {"ok": False, "error": str(e)}
        logger.warning(f"Risk-free rate fetch failed ({e}); using fallback {RISK_FREE:.4f}")

    _write_market_data_status(erp_status, risk_free_status)


def main():
    parser = argparse.ArgumentParser(
        description="HessGrp FCFF/FCFE valuation engine.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "examples:\n"
            "  python av_fcff_2.py --ticker AAPL\n"
            "  python av_fcff_2.py --index sp500 --limit 50\n"
            "  python av_fcff_2.py --filings ~/HessGrp/data/pending_valuations_2026-04-22.json\n"
            "  python av_fcff_2.py              # interactive menu"
        ),
    )
    parser.add_argument("--ticker",   metavar="SYM", action="append",
                        help="Value one or more stocks (repeat for multiple: --ticker AAPL --ticker JBL). Excel output + DB update.")
    parser.add_argument("--index",    choices=["sp500", "r2000"],
                        help="Batch valuation across S&P 500 or Russell 2000")
    parser.add_argument("--filings",  metavar="FILE",
                        help="Tickers from sec_daily_index JSON/xlsx/txt output")
    parser.add_argument("--limit",    type=int, default=None, metavar="N",
                        help="Cap batch to first N tickers")
    parser.add_argument("--growth",   type=int, default=GROWTH_PERIOD, metavar="N",
                        help=f"High-growth period in years (default {GROWTH_PERIOD})")
    parser.add_argument("--db",       default=DEFAULT_DB, metavar="PATH",
                        help="Path to valuation.db (default: $VALUATION_DB or /Volumes/Financial_Data/valuation.db)")
    parser.add_argument("--equity-override", type=float, default=None, metavar="DOLLARS",
                        help="Override shareholders' equity (in dollars) for all tickers in this run. "
                             "Use when AV balance sheet data is known to be incorrect (e.g., PPG Q1 2026). "
                             "Example: --equity-override 8104000000")
    parser.add_argument("--provider", choices=["av", "intrinio"], default="intrinio",
                        help="Fundamentals data source (default: intrinio, with automatic fallback "
                             "to AV per-ticker if an Intrinio fetch fails — see docs/decisions.md, "
                             "'Data provider: Intrinio becomes primary'). Pass --provider av to force "
                             "AV only, no fallback, e.g. for a manual AV-side comparison run.")
    args = parser.parse_args()

    growth_period = args.growth

    global EQUITY_OVERRIDE, DATA_PROVIDER
    if args.equity_override is not None:
        EQUITY_OVERRIDE = args.equity_override
        print(f"  equity override active: ${EQUITY_OVERRIDE:,.0f}")
    DATA_PROVIDER = args.provider
    if DATA_PROVIDER == "intrinio":
        print("  data provider: Intrinio (primary), AV fallback on per-ticker failure")
    else:
        print("  data provider: AV only (--provider av — no Intrinio fallback)")
    db_path       = args.db

    # ---- Fetch market reference data (deferred from module level) --------
    print("Fetching market reference data (ERP, risk-free rate)...")
    refresh_market_data()
    print(f"  ERP: {EQ_PREM:.4f}   Risk-free: {RISK_FREE:.4f}")

    # ---- Determine run mode ---------------------------------------------
    tickers      = []
    single_stock = False
    ticker       = None

    if args.ticker:
        ticker_list = [t.strip().upper() for t in args.ticker]
        if len(ticker_list) == 1:
            ticker       = ticker_list[0]
            index_label  = ticker
            single_stock = True
        else:
            tickers     = ticker_list
            index_label = "+".join(ticker_list[:3]) + ("..." if len(ticker_list) > 3 else "")
            print(f"\nTicker-list mode: valuing {len(tickers)} stocks")

    elif args.filings:
        if not os.path.exists(args.filings):
            print(f"Error: filings file not found: {args.filings}", file=sys.stderr)
            sys.exit(1)
        tickers     = get_tickers_from_filings(args.filings)
        index_label = "filings"
        print(f"\nFilings mode: valuing {len(tickers)} stocks from {args.filings}")

    elif args.index:
        if args.index == "sp500":
            tickers     = get_sp500_tickers()
            index_label = "sp500"
        else:
            tickers     = get_russell2000_tickers()
            index_label = "r2000"

    else:
        # Interactive fallback — used when launched from hess_menu
        print("\nSelect index to value:")
        print("  1. S&P 500")
        print("  2. Russell 2000")
        print("  3. Single stock")
        while True:
            choice = input("Choice [1/2/3]: ").strip()
            if choice == "1":
                tickers     = get_sp500_tickers()
                index_label = "sp500"
                break
            elif choice == "2":
                tickers     = get_russell2000_tickers()
                index_label = "r2000"
                break
            elif choice == "3":
                ticker       = input("Enter ticker symbol: ").strip().upper()
                index_label  = ticker
                single_stock = True
                break
            else:
                print("Please enter 1, 2, or 3.")

    # ---- Single-stock path: Excel output + DB update --------------------
    if single_stock:
        output_file = os.path.join(
            _log_dir,
            f"value_{index_label}_{date.today().strftime('%Y%m%d')}.xlsx",
        )
        detail = value_stock_detail(ticker, growth_period, db_path)
        if detail:
            generate_xlsx(detail, output_file)
            try:
                db_conn = sqlite3.connect(db_path, timeout=30)
                db_conn.execute("PRAGMA journal_mode=WAL")
                create_table(db_conn)
                insert_valuation(db_conn, _stock_value_from_detail(detail))
                db_conn.close()
                print(f"Valuation for {ticker} saved to database.")
                _rescore_tickers(db_path, [ticker])
            except Exception as e:
                logger.warning(f"DB write failed for {ticker}: {e}")
        else:
            try:
                industry = hg_dcflib.get_industry(ticker)
            except Exception:
                industry = ""
            print(f"Valuation failed for {ticker}.")
        return

    # ---- Batch path: Excel output ---------------------------------------
    if args.limit:
        tickers = tickers[:args.limit]

    # Skip permanently-excluded tickers before the first attempt, not just
    # before tomorrow's Iggy-level retry filtering — see get_excluded_tickers()
    # docstring and docs/known_errors.md 2026-07-23. Batch modes only: a
    # deliberate multi-ticker request (e.g. --ticker A --ticker B) still goes
    # through this filter since it's the same "many tickers, one run" pattern
    # this exists to protect; true single-stock mode is handled separately
    # above and is never filtered, so a one-off re-check of an excluded ticker
    # (exactly how the 2026-07-23 exclusions were themselves verified) still works.
    excluded = get_excluded_tickers()
    if excluded:
        already_excluded = [t for t in tickers if t.upper() in excluded]
        if already_excluded:
            print(
                f"  Skipping {len(already_excluded)} permanently-excluded ticker(s) "
                f"(see data/excluded_tickers.json): {', '.join(already_excluded)}"
            )
            tickers = [t for t in tickers if t.upper() not in excluded]

    print(
        f"Prefetching quotes for {len(tickers)} tickers (separate batch — "
        f"see docs/known_errors.md 2026-07-22 AV support guidance)..."
    )
    prefetch_quotes(tickers, MY_API_KEY)

    print(
        f"Valuing {len(tickers)} {index_label.upper()} stocks (growth period = {growth_period} years) ..."
    )

    # Optionally write to DB
    try:
        db_conn = sqlite3.connect(db_path, timeout=30)
        db_conn.execute("PRAGMA journal_mode=WAL")
        create_table(db_conn)
    except Exception:
        db_conn = None
        logger.warning("Database unavailable; skipping DB writes.")

    valuations = []
    valued_tickers = []
    failed_tickers = []
    total = len(tickers)
    bar_width = 40
    start_time = time.time()
    for idx, ticker in enumerate(tickers, 1):
        result = value_stock(ticker, growth_period, db_path)
        if result:
            valuations.append(result)
            valued_tickers.append(result.ticker)
            if db_conn:
                try:
                    insert_valuation(db_conn, result)
                except Exception as e:
                    logger.warning(f"DB insert failed for {ticker}: {e}")
        else:
            failed_tickers.append(ticker)

        filled = int(bar_width * idx / total)
        bar = "#" * filled + "-" * (bar_width - filled)
        elapsed = int(time.time() - start_time)
        h, rem = divmod(elapsed, 3600)
        m, s = divmod(rem, 60)
        print(f"\r  {idx}/{total} [{bar}] {h:02d}:{m:02d}:{s:02d}", end="", flush=True)

    print()  # newline after progress bar

    # Second pass: retry tickers that failed on the first pass, after one
    # deliberate cool-off. AV support confirmed (2026-07-22) that a longer
    # cool-off clears transient per-minute micro-throttles better than our
    # existing fast in-place retries (5s/15s) alone — every retry sequence in
    # the run that prompted this fix exhausted both in-place retries without
    # ever succeeding. Rather than stretching the in-place backoff to AV's
    # suggested 60-90s ceiling (which would multiply added runtime across
    # every failing ticker on the first pass), we pay one 60s cool-off ONCE
    # here and then reuse the same fast in-place retries for the smaller
    # failed-only subset — cheaper, and this is also exactly the workflow
    # TASK-109/114/115 already proved out manually (re-run the stale subset
    # after time has passed), just automated into a single run.
    if failed_tickers:
        print(
            f"\n{len(failed_tickers)} ticker(s) failed on first pass — "
            f"retrying after a 60s cool-off..."
        )
        time.sleep(60)
        retry_total = len(failed_tickers)
        for idx, ticker in enumerate(failed_tickers, 1):
            result = value_stock(ticker, growth_period, db_path)
            if result:
                valuations.append(result)
                valued_tickers.append(result.ticker)
                if db_conn:
                    try:
                        insert_valuation(db_conn, result)
                    except Exception as e:
                        logger.warning(f"DB insert failed for {ticker}: {e}")
            print(f"\r  retry {idx}/{retry_total}", end="", flush=True)
        print()
        still_failed = [t for t in failed_tickers if t not in valued_tickers]
        if still_failed:
            print(
                f"{len(still_failed)} ticker(s) still failed after retry pass: "
                f"{', '.join(still_failed)}"
            )

    if db_conn:
        db_conn.close()

    _rescore_tickers(db_path, valued_tickers)

    valuations.sort(key=lambda v: v.margin_of_safety, reverse=True)

    _index_display = {"sp500": "S&P 500", "r2000": "Russell 2000"}.get(
        index_label, index_label
    )
    output_file = os.path.join(
        _log_dir,
        f"value_{index_label}_{date.today().strftime('%Y%m%d')}.xlsx",
    )
    generate_summary_xlsx(valuations, output_file, _index_display)
    print(f"Done. {len(valuations)}/{len(tickers)} stocks valued successfully.")


if __name__ == "__main__":
    main()
